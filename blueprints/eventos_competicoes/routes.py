# -*- coding: utf-8 -*-
"""
Blueprint Eventos e Competições — inscrições com formulário, adesão por academia.
Visível em modo associação e academia.
"""
import json
import os
import uuid
from datetime import datetime, date
from flask import Blueprint, render_template, redirect, url_for, flash, request, session, Response, jsonify, send_file, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from config import get_db_connection
from utils.formularios_campos import CAMPOS_ALUNO_PADRAO, listar_campos_por_grupo, get_label

bp_eventos_competicoes = Blueprint("eventos_competicoes", __name__, url_prefix="/eventos-competicoes")

UPLOAD_ANEXOS = "eventos_anexos"
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'jpeg', 'png', 'gif', 'zip', 'rar', 'txt'}


def _allowed_file(filename):
    """Verifica se a extensão do arquivo é permitida."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _salvar_anexo_evento(file_storage, evento_id):
    """Salva anexo de evento em static/uploads/eventos_anexos/."""
    if not file_storage or not file_storage.filename:
        return None
    
    if not _allowed_file(file_storage.filename):
        return None
    
    ext = os.path.splitext(file_storage.filename)[1].lower()
    filename_original = secure_filename(file_storage.filename)
    filename_safe = f"evento_{evento_id}_{uuid.uuid4().hex[:12]}{ext}"
    
    folder = os.path.join(current_app.root_path, "static", "uploads", UPLOAD_ANEXOS)
    os.makedirs(folder, exist_ok=True)
    filepath = os.path.join(folder, filename_safe)
    
    try:
        file_storage.save(filepath)
        tamanho = os.path.getsize(filepath)
        return {
            "nome_original": filename_original,
            "caminho": filename_safe,
            "tamanho": tamanho,
            "tipo_mime": file_storage.content_type or "application/octet-stream"
        }
    except Exception as e:
        current_app.logger.error(f"Erro ao salvar anexo: {e}")
        return None


def _ids_academias_associacao(cursor, id_associacao):
    """Retorna ids das academias da associação."""
    cursor.execute("SELECT id FROM academias WHERE id_associacao = %s ORDER BY nome", (id_associacao,))
    return [r["id"] for r in cursor.fetchall()]


def _evento_encerrado(ev):
    """True se data_fim já passou."""
    if not ev or not ev.get("data_fim"):
        return False
    df = ev["data_fim"]
    if hasattr(df, "replace"):
        try:
            df = datetime.strptime(df[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            return False
    return datetime.now() > df


@bp_eventos_competicoes.route("/")
@login_required
def lista():
    """Lista eventos: associação vê os que criou; academia vê os disponíveis para aderir/gerenciar."""
    modo = session.get("modo_painel") or ""
    if modo not in ("associacao", "academia"):
        flash("Disponível apenas em modo associação ou academia.", "danger")
        return redirect(url_for("painel.home"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
        ids_acad = _get_ids_academias(cur)

        if modo == "associacao" and (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
            if not id_assoc:
                flash("Selecione a associação.", "warning")
                return redirect(url_for("associacao.gerenciamento_associacao"))
            # Verificar se colunas de taxa existem
            tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
            if tem_coluna_taxa:
                cur.execute("""
                    SELECT ec.id, ec.nome, ec.descricao, ec.tipo, ec.data_inicio, ec.data_fim, ec.status, ec.id_formulario, f.nome as formulario_nome,
                           COALESCE(ec.tem_taxa, 0) as tem_taxa, ec.valor_taxa_sugerido,
                           (SELECT COUNT(*) FROM eventos_competicoes_adesao WHERE evento_id=ec.id AND aderiu=1) as academias_aderidas,
                           (SELECT COUNT(*) FROM eventos_competicoes_inscricoes WHERE evento_id=ec.id AND status='enviada') as inscricoes_enviadas,
                           (SELECT COUNT(*) FROM eventos_competicoes_anexos WHERE evento_id=ec.id) as total_anexos
                    FROM eventos_competicoes ec
                    LEFT JOIN formularios f ON f.id = ec.id_formulario
                    WHERE ec.id_associacao = %s
                    ORDER BY ec.data_fim DESC
                """, (id_assoc,))
            else:
                cur.execute("""
                    SELECT ec.id, ec.nome, ec.descricao, ec.tipo, ec.data_inicio, ec.data_fim, ec.status, ec.id_formulario, f.nome as formulario_nome,
                           0 as tem_taxa, NULL as valor_taxa_sugerido,
                           (SELECT COUNT(*) FROM eventos_competicoes_adesao WHERE evento_id=ec.id AND aderiu=1) as academias_aderidas,
                           (SELECT COUNT(*) FROM eventos_competicoes_inscricoes WHERE evento_id=ec.id AND status='enviada') as inscricoes_enviadas,
                           (SELECT COUNT(*) FROM eventos_competicoes_anexos WHERE evento_id=ec.id) as total_anexos
                    FROM eventos_competicoes ec
                    LEFT JOIN formularios f ON f.id = ec.id_formulario
                    WHERE ec.id_associacao = %s
                    ORDER BY ec.data_fim DESC
                """, (id_assoc,))
            eventos = cur.fetchall()
            for ev in eventos:
                ev["encerrado"] = _evento_encerrado(ev)
                # Garantir que tem_taxa seja int para comparação no template
                if 'tem_taxa' in ev:
                    ev['tem_taxa'] = int(ev['tem_taxa']) if ev['tem_taxa'] else 0
                # Buscar anexos do evento
                cur.execute("""
                    SELECT id, nome_arquivo, tamanho_bytes, descricao
                    FROM eventos_competicoes_anexos
                    WHERE evento_id = %s
                    ORDER BY created_at DESC
                """, (ev["id"],))
                anexos = cur.fetchall()
                # Formatar tamanhos
                for anexo in anexos:
                    tamanho = anexo.get("tamanho_bytes") or 0
                    if tamanho < 1024:
                        anexo["tamanho_formatado"] = f"{tamanho} B"
                    elif tamanho < 1024 * 1024:
                        anexo["tamanho_formatado"] = f"{tamanho / 1024:.1f} KB"
                    else:
                        anexo["tamanho_formatado"] = f"{tamanho / (1024 * 1024):.1f} MB"
                ev["anexos"] = anexos
            return render_template("eventos_competicoes/lista_associacao.html",
                eventos=eventos, back_url=url_for("associacao.gerenciamento_associacao"))

        if modo == "academia" and (current_user.has_role("gestor_academia") or current_user.has_role("professor") or current_user.has_role("admin")):
            if not ids_acad:
                flash("Nenhuma academia vinculada.", "warning")
                return redirect(url_for("painel.home"))
            academia_id = request.args.get("academia_id", type=int) or (ids_acad[0] if ids_acad else None)
            if academia_id not in ids_acad:
                academia_id = ids_acad[0]
            # Verificar se colunas de taxa existem
            tem_coluna_taxa, tem_coluna_valor_taxa = _colunas_taxa_existem(cur)
            if tem_coluna_taxa and tem_coluna_valor_taxa:
                cur.execute("""
                    SELECT ec.id, ec.nome, ec.descricao, ec.tipo, ec.data_inicio, ec.data_fim, ec.id_formulario, f.nome as formulario_nome,
                           COALESCE(ec.tem_taxa, 0) as tem_taxa, ec.valor_taxa_sugerido,
                           COALESCE(ea.aderiu, 0) as aderiu, ea.academia_id, ea.valor_taxa
                    FROM eventos_competicoes ec
                    INNER JOIN academias ac ON ac.id_associacao = ec.id_associacao AND ac.id = %s
                    LEFT JOIN formularios f ON f.id = ec.id_formulario
                    LEFT JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s
                    ORDER BY ec.data_fim DESC
                """, (academia_id, academia_id))
            else:
                cur.execute("""
                    SELECT ec.id, ec.nome, ec.descricao, ec.tipo, ec.data_inicio, ec.data_fim, ec.id_formulario, f.nome as formulario_nome,
                           0 as tem_taxa, NULL as valor_taxa_sugerido,
                           COALESCE(ea.aderiu, 0) as aderiu, ea.academia_id, NULL as valor_taxa
                    FROM eventos_competicoes ec
                    INNER JOIN academias ac ON ac.id_associacao = ec.id_associacao AND ac.id = %s
                    LEFT JOIN formularios f ON f.id = ec.id_formulario
                    LEFT JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s
                    ORDER BY ec.data_fim DESC
                """, (academia_id, academia_id))
            eventos = cur.fetchall()
            for ev in eventos:
                ev["encerrado"] = _evento_encerrado(ev)
                # Garantir que tem_taxa seja int para comparação no template
                if 'tem_taxa' in ev:
                    ev['tem_taxa'] = int(ev['tem_taxa']) if ev['tem_taxa'] else 0
                # Buscar anexos do evento
                cur.execute("""
                    SELECT id, nome_arquivo, tamanho_bytes, descricao
                    FROM eventos_competicoes_anexos
                    WHERE evento_id = %s
                    ORDER BY created_at DESC
                """, (ev["id"],))
                anexos = cur.fetchall()
                # Formatar tamanhos
                for anexo in anexos:
                    tamanho = anexo.get("tamanho_bytes") or 0
                    if tamanho < 1024:
                        anexo["tamanho_formatado"] = f"{tamanho} B"
                    elif tamanho < 1024 * 1024:
                        anexo["tamanho_formatado"] = f"{tamanho / 1024:.1f} KB"
                    else:
                        anexo["tamanho_formatado"] = f"{tamanho / (1024 * 1024):.1f} MB"
                ev["anexos"] = anexos
            cur.execute("SELECT id, nome FROM academias WHERE id IN (%s) ORDER BY nome" % ",".join(["%s"] * len(ids_acad)), tuple(ids_acad))
            academias_sel = cur.fetchall()
            return render_template("eventos_competicoes/lista_academia.html",
                eventos=eventos, academias=academias_sel, academias_ids=ids_acad, academia_id=academia_id,
                back_url=url_for("academia.painel_academia", academia_id=academia_id) if academia_id else url_for("painel.home"))

        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    finally:
        cur.close()
        conn.close()


def _colunas_taxa_existem(cur):
    """Verifica se as colunas de taxa existem nas tabelas."""
    try:
        cur.execute("SHOW COLUMNS FROM eventos_competicoes LIKE 'tem_taxa'")
        tem_taxa_exists = cur.fetchone() is not None
        cur.execute("SHOW COLUMNS FROM eventos_competicoes_adesao LIKE 'valor_taxa'")
        valor_taxa_exists = cur.fetchone() is not None
        return tem_taxa_exists, valor_taxa_exists
    except Exception:
        return False, False


def _criar_pagamento_inscricao(cur, inscricao_id, evento_id, academia_id):
    """Cria registro de pagamento para uma inscrição se o evento tiver taxa."""
    try:
        # Verificar se tabela de pagamentos existe
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
            if not cur.fetchone():
                return False
        except Exception:
            return False
        
        # Verificar se colunas de taxa existem
        tem_coluna_taxa, tem_coluna_valor_taxa = _colunas_taxa_existem(cur)
        if not tem_coluna_taxa or not tem_coluna_valor_taxa:
            return False
        
        # Buscar informações do evento e adesão
        cur.execute("""
            SELECT COALESCE(ec.tem_taxa, 0) as tem_taxa, ea.valor_taxa
            FROM eventos_competicoes ec
            LEFT JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s
            WHERE ec.id = %s
        """, (academia_id, evento_id))
        ev_info = cur.fetchone()
        
        if not ev_info:
            return False
        
        # Verificar se evento tem taxa (tem_taxa == 1) e se tem valor definido
        tem_taxa = ev_info.get("tem_taxa") if ev_info else 0
        tem_taxa_int = int(tem_taxa) if tem_taxa else 0
        
        if tem_taxa_int == 1 and ev_info.get("valor_taxa"):
            valor = float(ev_info.get("valor_taxa"))
            if valor > 0:
                # Verificar se já existe registro de pagamento
                cur.execute("SELECT id FROM eventos_competicoes_inscricoes_pagamentos WHERE inscricao_id = %s", (inscricao_id,))
                if not cur.fetchone():
                    cur.execute("""
                        INSERT INTO eventos_competicoes_inscricoes_pagamentos (inscricao_id, valor)
                        VALUES (%s, %s)
                    """, (inscricao_id, valor))
                    return True
                return True
    except Exception as e:
        # Log erro mas não interrompe o fluxo
        import traceback
        print(f"Erro ao criar pagamento para inscrição {inscricao_id}: {e}")
        print(traceback.format_exc())
    return False


def _get_ids_academias(cur):
    """Ids das academias do usuário (modo academia)."""
    ids = []
    # Primeiro: verificar usuarios_academias (prioridade)
    cur.execute("""
        SELECT academia_id FROM usuarios_academias WHERE usuario_id = %s ORDER BY academia_id
    """, (current_user.id,))
    vinculadas = [r["academia_id"] for r in cur.fetchall()]
    if vinculadas:
        return vinculadas
    
    # Modo academia: gestor_academia/professor só veem academias de usuarios_academias (não id_academia)
    if session.get("modo_painel") == "academia" and (current_user.has_role("gestor_academia") or current_user.has_role("professor")):
        return []
    
    # Admin: todas as academias
    if current_user.has_role("admin"):
        cur.execute("SELECT id FROM academias ORDER BY nome")
        ids = [r["id"] for r in cur.fetchall()]
        return ids
    
    # Gestor federação: academias da federação
    if current_user.has_role("gestor_federacao"):
        cur.execute("""
            SELECT ac.id FROM academias ac 
            JOIN associacoes ass ON ass.id = ac.id_associacao 
            WHERE ass.id_federacao = %s ORDER BY ac.nome
        """, (getattr(current_user, "id_federacao", None),))
        ids = [r["id"] for r in cur.fetchall()]
        return ids
    
    # Gestor associação: academias da associação
    if current_user.has_role("gestor_associacao"):
        id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
        if id_assoc:
            cur.execute("SELECT id FROM academias WHERE id_associacao = %s ORDER BY nome", (id_assoc,))
            ids = [r["id"] for r in cur.fetchall()]
        return ids
    
    # Fallback: id_academia (se existir)
    if getattr(current_user, "id_academia", None):
        return [current_user.id_academia]
    
    return []


@bp_eventos_competicoes.route("/cadastro", methods=["GET", "POST"])
@login_required
def cadastro():
    """Associação cria evento com formulário e data fim."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))

    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, nome FROM formularios WHERE id_associacao = %s AND ativo = 1 ORDER BY nome", (id_assoc,))
        formularios = cur.fetchall()

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            descricao = request.form.get("descricao", "").strip()
            tipo = request.form.get("tipo", "evento")
            id_formulario = request.form.get("id_formulario", type=int) or None
            data_inicio = request.form.get("data_inicio") or None
            data_fim = request.form.get("data_fim")
            if not nome or not data_fim:
                flash("Informe nome e data/hora fim do evento.", "danger")
                return render_template("eventos_competicoes/cadastro.html", formularios=formularios,
                    back_url=url_for("eventos_competicoes.lista"))
            if tipo not in ("evento", "competicao"):
                tipo = "evento"
            try:
                dt_fim = datetime.strptime(data_fim[:16], "%Y-%m-%dT%H:%M") if data_fim else None
            except Exception:
                dt_fim = None
            if not dt_fim:
                flash("Data/hora fim inválida.", "danger")
                return render_template("eventos_competicoes/cadastro.html", formularios=formularios,
                    back_url=url_for("eventos_competicoes.lista"))

            dt_ini = None
            if data_inicio:
                try:
                    dt_ini = datetime.strptime(data_inicio[:16], "%Y-%m-%dT%H:%M")
                except Exception:
                    pass

            # Campos de taxa (verificar se colunas existem)
            tem_taxa = request.form.get("tem_taxa") == "1"
            valor_taxa_sugerido = None
            if tem_taxa:
                valor_str = request.form.get("valor_taxa_sugerido", "").strip()
                if not valor_str:
                    flash("Quando o evento tem taxa, o valor sugerido é obrigatório.", "danger")
                    return render_template("eventos_competicoes/cadastro.html", formularios=formularios,
                        back_url=url_for("eventos_competicoes.lista"))
                try:
                    valor_taxa_sugerido = float(valor_str)
                    if valor_taxa_sugerido < 0:
                        flash("O valor da taxa deve ser positivo.", "danger")
                        return render_template("eventos_competicoes/cadastro.html", formularios=formularios,
                            back_url=url_for("eventos_competicoes.lista"))
                except (ValueError, TypeError):
                    flash("Valor da taxa inválido.", "danger")
                    return render_template("eventos_competicoes/cadastro.html", formularios=formularios,
                        back_url=url_for("eventos_competicoes.lista"))

            # Verificar se colunas de taxa existem
            tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
            
            # Se as colunas não existem, tentar criar automaticamente
            if not tem_coluna_taxa:
                try:
                    cur.execute("ALTER TABLE eventos_competicoes ADD COLUMN tem_taxa TINYINT(1) NOT NULL DEFAULT 0")
                    cur.execute("ALTER TABLE eventos_competicoes ADD COLUMN valor_taxa_sugerido DECIMAL(10,2) NULL DEFAULT NULL")
                    conn.commit()  # Commit para criar as colunas
                    tem_coluna_taxa = True
                except Exception as e:
                    # Se já existem, ignorar erro
                    if "Duplicate column name" not in str(e):
                        try:
                            current_app.logger.error(f"Erro ao criar colunas: {e}")
                        except:
                            pass
                    # Verificar novamente
                    tem_coluna_taxa, _ = _colunas_taxa_existem(cur)

            # Converter tem_taxa para int (0 ou 1) para MySQL TINYINT
            tem_taxa_int = 1 if tem_taxa else 0
            # Garantir que valor_taxa_sugerido seja None quando tem_taxa é False
            if not tem_taxa:
                valor_taxa_sugerido = None
            
            # Executar INSERT
            if tem_coluna_taxa:
                try:
                    cur.execute("""
                        INSERT INTO eventos_competicoes (nome, descricao, id_associacao, id_formulario, tipo, data_inicio, data_fim, tem_taxa, valor_taxa_sugerido)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (nome, descricao or None, id_assoc, id_formulario, tipo, dt_ini, dt_fim, tem_taxa_int, valor_taxa_sugerido))
                except Exception as e:
                    try:
                        current_app.logger.error(f"Erro ao criar evento com taxa: {e}", exc_info=True)
                    except:
                        pass
                    flash(f"Erro ao salvar dados da taxa: {str(e)}", "danger")
                    return render_template("eventos_competicoes/cadastro.html", formularios=formularios,
                        back_url=url_for("eventos_competicoes.lista"))
            else:
                cur.execute("""
                    INSERT INTO eventos_competicoes (nome, descricao, id_associacao, id_formulario, tipo, data_inicio, data_fim)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (nome, descricao or None, id_assoc, id_formulario, tipo, dt_ini, dt_fim))
            evento_id = cur.lastrowid
            
            # Processar anexos enviados
            anexos_enviados = request.files.getlist("anexos")
            for anexo_file in anexos_enviados:
                if anexo_file and anexo_file.filename:
                    resultado = _salvar_anexo_evento(anexo_file, evento_id)
                    if resultado:
                        descricao_anexo = request.form.get(f"descricao_anexo_{anexo_file.filename}", "").strip() or None
                        cur.execute("""
                            INSERT INTO eventos_competicoes_anexos 
                            (evento_id, nome_arquivo, caminho_arquivo, tamanho_bytes, tipo_mime, descricao, criado_por)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (
                            evento_id,
                            resultado["nome_original"],
                            resultado["caminho"],
                            resultado["tamanho"],
                            resultado["tipo_mime"],
                            descricao_anexo,
                            current_user.id
                        ))

            academias_ids = _ids_academias_associacao(cur, id_assoc)
            for ac_id in academias_ids:
                cur.execute("""
                    INSERT INTO eventos_competicoes_adesao (evento_id, academia_id, aderiu) VALUES (%s, %s, 0)
                """, (evento_id, ac_id))

            # Inserir no calendário da associação e das academias automaticamente
            dt_ini_date = dt_ini.date() if dt_ini else (dt_fim.date() if dt_fim else date.today())
            dt_fim_date = dt_fim.date() if dt_fim else dt_ini_date
            hora_ini = dt_ini.time() if dt_ini and hasattr(dt_ini, "time") else None
            hora_fim = dt_fim.time() if dt_fim and hasattr(dt_fim, "time") else None
            tipo_cal = "competicao" if tipo == "competicao" else "evento"
            try:
                cur.execute("""
                    INSERT INTO eventos (titulo, descricao, data_inicio, data_fim, hora_inicio, hora_fim, tipo, recorrente,
                        nivel, nivel_id, criado_por_usuario_id, origem_sincronizacao, evento_competicao_id, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 0, 'associacao', %s, %s, 'eventos_competicoes', %s, 'ativo')
                """, (nome, descricao or None, dt_ini_date, dt_fim_date, hora_ini, hora_fim, tipo_cal, id_assoc, current_user.id, evento_id))
                for ac_id in academias_ids:
                    cur.execute("""
                        INSERT INTO eventos (titulo, descricao, data_inicio, data_fim, hora_inicio, hora_fim, tipo, recorrente,
                            nivel, nivel_id, criado_por_usuario_id, origem_sincronizacao, evento_competicao_id, status)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, 0, 'academia', %s, %s, 'eventos_competicoes', %s, 'ativo')
                    """, (nome, descricao or None, dt_ini_date, dt_fim_date, hora_ini, hora_fim, tipo_cal, ac_id, current_user.id, evento_id))
            except Exception:
                try:
                    cur.execute("""
                        INSERT INTO eventos (titulo, descricao, data_inicio, data_fim, hora_inicio, hora_fim, tipo, recorrente,
                            nivel, nivel_id, criado_por_usuario_id, origem_sincronizacao, status)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, 0, 'associacao', %s, %s, 'eventos_competicoes', 'ativo')
                    """, (nome, descricao or None, dt_ini_date, dt_fim_date, hora_ini, hora_fim, tipo_cal, id_assoc, current_user.id))
                    for ac_id in academias_ids:
                        cur.execute("""
                            INSERT INTO eventos (titulo, descricao, data_inicio, data_fim, hora_inicio, hora_fim, tipo, recorrente,
                                nivel, nivel_id, criado_por_usuario_id, origem_sincronizacao, status)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, 0, 'academia', %s, %s, 'eventos_competicoes', 'ativo')
                        """, (nome, descricao or None, dt_ini_date, dt_fim_date, hora_ini, hora_fim, tipo_cal, ac_id, current_user.id))
                except Exception:
                    pass

            conn.commit()
            flash("Evento criado. As academias podem aderir ao evento.", "success")
            return redirect(url_for("eventos_competicoes.lista"))

        # Buscar anexos existentes (vazio no cadastro)
        anexos_existentes = []
        
        return render_template("eventos_competicoes/cadastro.html", formularios=formularios,
            anexos_existentes=anexos_existentes, back_url=url_for("eventos_competicoes.lista"))
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/editar/<int:evento_id>", methods=["GET", "POST"])
@login_required
def editar(evento_id):
    """Associação edita evento existente."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))

    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se colunas de taxa existem
        tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
        if tem_coluna_taxa:
            cur.execute("""
                SELECT id, nome, descricao, tipo, id_formulario, data_inicio, data_fim, 
                       COALESCE(tem_taxa, 0) as tem_taxa, valor_taxa_sugerido
                FROM eventos_competicoes WHERE id = %s AND id_associacao = %s
            """, (evento_id, id_assoc))
        else:
            cur.execute("""
                SELECT id, nome, descricao, tipo, id_formulario, data_inicio, data_fim, 0 as tem_taxa, NULL as valor_taxa_sugerido
                FROM eventos_competicoes WHERE id = %s AND id_associacao = %s
            """, (evento_id, id_assoc))
        evento = cur.fetchone()
        if not evento:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))
        # Garantir que tem_taxa seja int
        if 'tem_taxa' in evento:
            evento['tem_taxa'] = int(evento['tem_taxa']) if evento['tem_taxa'] else 0
        if not evento:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))

        cur.execute("SELECT id, nome FROM formularios WHERE id_associacao = %s AND ativo = 1 ORDER BY nome", (id_assoc,))
        formularios = cur.fetchall()
        
        # Buscar anexos existentes
        cur.execute("""
            SELECT id, nome_arquivo, caminho_arquivo, tamanho_bytes, tipo_mime, descricao, created_at
            FROM eventos_competicoes_anexos
            WHERE evento_id = %s
            ORDER BY created_at DESC
        """, (evento_id,))
        anexos_existentes = cur.fetchall()
        for anexo in anexos_existentes:
            # Formatar tamanho
            tamanho = anexo.get("tamanho_bytes") or 0
            if tamanho < 1024:
                anexo["tamanho_formatado"] = f"{tamanho} B"
            elif tamanho < 1024 * 1024:
                anexo["tamanho_formatado"] = f"{tamanho / 1024:.1f} KB"
            else:
                anexo["tamanho_formatado"] = f"{tamanho / (1024 * 1024):.1f} MB"

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            descricao = request.form.get("descricao", "").strip()
            tipo = request.form.get("tipo", "evento")
            id_formulario = request.form.get("id_formulario", type=int) or None
            data_inicio = request.form.get("data_inicio") or None
            data_fim = request.form.get("data_fim")
            if not nome or not data_fim:
                flash("Informe nome e data/hora fim do evento.", "danger")
                return render_template("eventos_competicoes/editar.html", evento=evento, formularios=formularios,
                    anexos_existentes=anexos_existentes, back_url=url_for("eventos_competicoes.lista"))
            if tipo not in ("evento", "competicao"):
                tipo = "evento"
            try:
                dt_fim = datetime.strptime(data_fim[:16], "%Y-%m-%dT%H:%M") if data_fim else None
            except Exception:
                dt_fim = None
            if not dt_fim:
                flash("Data/hora fim inválida.", "danger")
                return render_template("eventos_competicoes/editar.html", evento=evento, formularios=formularios,
                    anexos_existentes=anexos_existentes, back_url=url_for("eventos_competicoes.lista"))
            
            # Remover anexos marcados para exclusão
            anexos_remover = request.form.getlist("remover_anexo")
            for anexo_id in anexos_remover:
                try:
                    anexo_id_int = int(anexo_id)
                    # Buscar caminho do arquivo antes de deletar
                    cur.execute("SELECT caminho_arquivo FROM eventos_competicoes_anexos WHERE id = %s AND evento_id = %s", 
                               (anexo_id_int, evento_id))
                    anexo_row = cur.fetchone()
                    if anexo_row:
                        # Deletar arquivo físico
                        filepath = os.path.join(current_app.root_path, "static", "uploads", UPLOAD_ANEXOS, anexo_row["caminho_arquivo"])
                        try:
                            if os.path.exists(filepath):
                                os.remove(filepath)
                        except Exception:
                            pass
                        # Deletar registro no banco
                        cur.execute("DELETE FROM eventos_competicoes_anexos WHERE id = %s AND evento_id = %s", 
                                   (anexo_id_int, evento_id))
                except Exception:
                    pass

            dt_ini = None
            if data_inicio:
                try:
                    dt_ini = datetime.strptime(data_inicio[:16], "%Y-%m-%dT%H:%M")
                except Exception:
                    pass

            # Campos de taxa (verificar se colunas existem)
            tem_taxa = request.form.get("tem_taxa") == "1"
            valor_taxa_sugerido = None
            if tem_taxa:
                valor_str = request.form.get("valor_taxa_sugerido", "").strip()
                if not valor_str:
                    flash("Quando o evento tem taxa, o valor sugerido é obrigatório.", "danger")
                    return render_template("eventos_competicoes/editar.html", evento=evento, formularios=formularios,
                        anexos_existentes=anexos_existentes, back_url=url_for("eventos_competicoes.lista"))
                try:
                    valor_taxa_sugerido = float(valor_str)
                    if valor_taxa_sugerido < 0:
                        flash("O valor da taxa deve ser positivo.", "danger")
                        return render_template("eventos_competicoes/editar.html", evento=evento, formularios=formularios,
                            anexos_existentes=anexos_existentes, back_url=url_for("eventos_competicoes.lista"))
                except (ValueError, TypeError):
                    flash("Valor da taxa inválido.", "danger")
                    return render_template("eventos_competicoes/editar.html", evento=evento, formularios=formularios,
                        anexos_existentes=anexos_existentes, back_url=url_for("eventos_competicoes.lista"))

            # Verificar se colunas de taxa existem
            tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
            
            # Se as colunas não existem, tentar criar automaticamente
            if not tem_coluna_taxa:
                try:
                    cur.execute("ALTER TABLE eventos_competicoes ADD COLUMN tem_taxa TINYINT(1) NOT NULL DEFAULT 0")
                    cur.execute("ALTER TABLE eventos_competicoes ADD COLUMN valor_taxa_sugerido DECIMAL(10,2) NULL DEFAULT NULL")
                    conn.commit()  # Commit para criar as colunas
                    tem_coluna_taxa = True
                    try:
                        current_app.logger.info("Colunas de taxa criadas automaticamente")
                    except:
                        pass
                except Exception as e:
                    # Se já existem, ignorar erro
                    try:
                        if "Duplicate column name" not in str(e):
                            current_app.logger.error(f"Erro ao criar colunas: {e}")
                    except:
                        pass
                    # Verificar novamente
                    tem_coluna_taxa, _ = _colunas_taxa_existem(cur)

            # Converter tem_taxa para int (0 ou 1) para MySQL TINYINT
            tem_taxa_int = 1 if tem_taxa else 0
            # Garantir que valor_taxa_sugerido seja None quando tem_taxa é False
            if not tem_taxa:
                valor_taxa_sugerido = None
            
            # Executar UPDATE
            if tem_coluna_taxa:
                try:
                    cur.execute("""
                        UPDATE eventos_competicoes
                        SET nome = %s, descricao = %s, tipo = %s, id_formulario = %s, data_inicio = %s, data_fim = %s, tem_taxa = %s, valor_taxa_sugerido = %s
                        WHERE id = %s AND id_associacao = %s
                    """, (nome, descricao or None, tipo, id_formulario, dt_ini, dt_fim, tem_taxa_int, valor_taxa_sugerido, evento_id, id_assoc))
                    # Verificar se a atualização foi bem-sucedida
                    if cur.rowcount == 0:
                        flash("Erro ao atualizar evento. Verifique se o evento existe.", "danger")
                        return render_template("eventos_competicoes/editar.html", evento=evento, formularios=formularios,
                            anexos_existentes=anexos_existentes, back_url=url_for("eventos_competicoes.lista"))
                except Exception as e:
                    try:
                        current_app.logger.error(f"Erro ao atualizar evento com taxa: {e}", exc_info=True)
                    except:
                        pass
                    flash(f"Erro ao salvar dados da taxa: {str(e)}", "danger")
                    return render_template("eventos_competicoes/editar.html", evento=evento, formularios=formularios,
                        anexos_existentes=anexos_existentes, back_url=url_for("eventos_competicoes.lista"))
            else:
                # Se ainda não tem colunas, salvar sem elas
                cur.execute("""
                    UPDATE eventos_competicoes
                    SET nome = %s, descricao = %s, tipo = %s, id_formulario = %s, data_inicio = %s, data_fim = %s
                    WHERE id = %s AND id_associacao = %s
                """, (nome, descricao or None, tipo, id_formulario, dt_ini, dt_fim, evento_id, id_assoc))
            
            # Processar novos anexos enviados
            anexos_enviados = request.files.getlist("anexos")
            for anexo_file in anexos_enviados:
                if anexo_file and anexo_file.filename:
                    resultado = _salvar_anexo_evento(anexo_file, evento_id)
                    if resultado:
                        descricao_anexo = request.form.get(f"descricao_anexo_{anexo_file.filename}", "").strip() or None
                        cur.execute("""
                            INSERT INTO eventos_competicoes_anexos 
                            (evento_id, nome_arquivo, caminho_arquivo, tamanho_bytes, tipo_mime, descricao, criado_por)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (
                            evento_id,
                            resultado["nome_original"],
                            resultado["caminho"],
                            resultado["tamanho"],
                            resultado["tipo_mime"],
                            descricao_anexo,
                            current_user.id
                        ))

            # Atualizar eventos do calendário vinculados
            dt_ini_date = dt_ini.date() if dt_ini else (dt_fim.date() if dt_fim else date.today())
            dt_fim_date = dt_fim.date() if dt_fim else dt_ini_date
            hora_ini = dt_ini.time() if dt_ini and hasattr(dt_ini, "time") else None
            hora_fim = dt_fim.time() if dt_fim and hasattr(dt_fim, "time") else None
            tipo_cal = "competicao" if tipo == "competicao" else "evento"
            try:
                cur.execute("""
                    UPDATE eventos
                    SET titulo = %s, descricao = %s, data_inicio = %s, data_fim = %s, hora_inicio = %s, hora_fim = %s, tipo = %s
                    WHERE evento_competicao_id = %s
                """, (nome, descricao or None, dt_ini_date, dt_fim_date, hora_ini, hora_fim, tipo_cal, evento_id))
            except Exception:
                pass

            conn.commit()
            flash("Evento atualizado.", "success")
            return redirect(url_for("eventos_competicoes.lista"))

        # GET: formatar datas para datetime-local
        data_inicio_val = ""
        data_fim_val = ""
        if evento.get("data_inicio"):
            di = evento["data_inicio"]
            if hasattr(di, "strftime"):
                data_inicio_val = di.strftime("%Y-%m-%dT%H:%M")
            else:
                data_inicio_val = str(di)[:16].replace(" ", "T")
        if evento.get("data_fim"):
            df = evento["data_fim"]
            if hasattr(df, "strftime"):
                data_fim_val = df.strftime("%Y-%m-%dT%H:%M")
            else:
                data_fim_val = str(df)[:16].replace(" ", "T")

        return render_template("eventos_competicoes/editar.html",
            evento=evento, formularios=formularios,
            data_inicio_val=data_inicio_val, data_fim_val=data_fim_val,
            back_url=url_for("eventos_competicoes.lista"))
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/finalizar/<int:evento_id>", methods=["POST"])
@login_required
def finalizar(evento_id):
    """Associação marca evento como finalizado ou reativa."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))

    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        return redirect(url_for("associacao.gerenciamento_associacao"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, status FROM eventos_competicoes WHERE id = %s AND id_associacao = %s", (evento_id, id_assoc))
        ev = cur.fetchone()
        if not ev:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))

        novo_status = "finalizado" if ev.get("status") != "finalizado" else "ativo"
        cur.execute("UPDATE eventos_competicoes SET status = %s WHERE id = %s", (novo_status, evento_id))
        conn.commit()
        if novo_status == "finalizado":
            flash("Evento finalizado.", "success")
        else:
            flash("Evento reativado.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.lista"))


@bp_eventos_competicoes.route("/aderir/<int:evento_id>", methods=["GET", "POST"])
@login_required
def aderir(evento_id):
    """Academia adere ao evento (aderiu=1). Se evento tem taxa, mostra formulário para definir valor."""
    academia_id = request.args.get("academia_id", type=int) or request.form.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.lista", academia_id=academia_id))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se colunas de taxa existem
        tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
        if tem_coluna_taxa:
            cur.execute("""
                SELECT id, nome, data_fim, tem_taxa, valor_taxa_sugerido
                FROM eventos_competicoes WHERE id = %s
            """, (evento_id,))
        else:
            cur.execute("""
                SELECT id, nome, data_fim, 0 as tem_taxa, NULL as valor_taxa_sugerido
                FROM eventos_competicoes WHERE id = %s
            """, (evento_id,))
        ev = cur.fetchone()
        if not ev:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista", academia_id=academia_id))
        if _evento_encerrado(ev):
            flash("Evento encerrado. Não é possível alterar adesão.", "warning")
            return redirect(url_for("eventos_competicoes.lista", academia_id=academia_id))
        
        # Verificar se já aderiu e qual valor está definido
        _, tem_coluna_valor_taxa = _colunas_taxa_existem(cur)
        if tem_coluna_valor_taxa:
            cur.execute("""
                SELECT aderiu, valor_taxa FROM eventos_competicoes_adesao 
                WHERE evento_id = %s AND academia_id = %s
            """, (evento_id, academia_id))
        else:
            cur.execute("""
                SELECT aderiu, NULL as valor_taxa FROM eventos_competicoes_adesao 
                WHERE evento_id = %s AND academia_id = %s
            """, (evento_id, academia_id))
        adesao_existente = cur.fetchone()
        
        if request.method == "GET":
            # Se evento tem taxa, sempre mostrar formulário para definir/alterar valor
            tem_taxa = ev.get("tem_taxa")
            # Verificar se tem_taxa é True ou 1 (pode vir como int ou bool)
            if tem_taxa and (tem_taxa == 1 or tem_taxa is True):
                return render_template("eventos_competicoes/aderir_com_taxa.html",
                    evento=ev, academia_id=academia_id,
                    valor_sugerido=ev.get("valor_taxa_sugerido"),
                    valor_atual=adesao_existente.get("valor_taxa") if adesao_existente else None,
                    ja_aderiu=adesao_existente.get("aderiu") if adesao_existente else False,
                    back_url=url_for("eventos_competicoes.lista", academia_id=academia_id))
            else:
                # Se não tem taxa, aderir diretamente
                return _toggle_adesao(evento_id, 1, academia_id, None)
        
        # POST: Processar adesão com valor da taxa
        valor_taxa = None
        tem_taxa = ev.get("tem_taxa")
        if tem_taxa and (tem_taxa == 1 or tem_taxa is True):
            try:
                valor_str = request.form.get("valor_taxa", "").strip()
                if valor_str:
                    valor_taxa = float(valor_str)
                    if valor_taxa < 0:
                        flash("Valor da taxa deve ser positivo.", "danger")
                        return render_template("eventos_competicoes/aderir_com_taxa.html",
                            evento=ev, academia_id=academia_id,
                            valor_sugerido=ev.get("valor_taxa_sugerido"),
                            valor_atual=adesao_existente.get("valor_taxa") if adesao_existente else None,
                            ja_aderiu=adesao_existente.get("aderiu") if adesao_existente else False,
                            back_url=url_for("eventos_competicoes.lista", academia_id=academia_id))
                    # Validar que não é menor que o sugerido (pode ser igual ou maior)
                    if ev.get("valor_taxa_sugerido") and valor_taxa < ev.get("valor_taxa_sugerido"):
                        flash(f"O valor deve ser igual ou maior que o sugerido (R$ {ev.get('valor_taxa_sugerido'):.2f}).", "danger")
                        return render_template("eventos_competicoes/aderir_com_taxa.html",
                            evento=ev, academia_id=academia_id,
                            valor_sugerido=ev.get("valor_taxa_sugerido"),
                            valor_atual=adesao_existente.get("valor_taxa") if adesao_existente else None,
                            ja_aderiu=adesao_existente.get("aderiu") if adesao_existente else False,
                            back_url=url_for("eventos_competicoes.lista", academia_id=academia_id))
            except (ValueError, TypeError):
                flash("Valor da taxa inválido.", "danger")
                return render_template("eventos_competicoes/aderir_com_taxa.html",
                    evento=ev, academia_id=academia_id,
                    valor_sugerido=ev.get("valor_taxa_sugerido"),
                    valor_atual=adesao_existente.get("valor_taxa") if adesao_existente else None,
                    ja_aderiu=adesao_existente.get("aderiu") if adesao_existente else False,
                    back_url=url_for("eventos_competicoes.lista", academia_id=academia_id))
        
        return _toggle_adesao(evento_id, 1, academia_id, valor_taxa)
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/desaderir/<int:evento_id>", methods=["POST"])
@login_required
def desaderir(evento_id):
    """Academia desiste da adesão (aderiu=0)."""
    academia_id = request.form.get("academia_id", type=int) or request.args.get("academia_id", type=int)
    return _toggle_adesao(evento_id, 0, academia_id, None)


def _toggle_adesao(evento_id, aderiu, academia_id, valor_taxa=None):
    """Atualiza adesão da academia ao evento. Se valor_taxa for fornecido, atualiza também."""
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.lista"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, data_fim FROM eventos_competicoes WHERE id = %s", (evento_id,))
        ev = cur.fetchone()
        if not ev:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista", academia_id=academia_id))
        if _evento_encerrado(ev):
            flash("Evento encerrado. Não é possível alterar adesão.", "warning")
            return redirect(url_for("eventos_competicoes.lista", academia_id=academia_id))

        # Verificar se coluna valor_taxa existe
        _, tem_coluna_valor_taxa = _colunas_taxa_existem(cur)
        
        # Se está aderindo e tem valor_taxa, atualizar ou inserir com valor
        if aderiu and valor_taxa is not None and tem_coluna_valor_taxa:
            cur.execute("""
                UPDATE eventos_competicoes_adesao 
                SET aderiu = %s, valor_taxa = %s 
                WHERE evento_id = %s AND academia_id = %s
            """, (aderiu, valor_taxa, evento_id, academia_id))
            if cur.rowcount == 0:
                cur.execute("""
                    INSERT INTO eventos_competicoes_adesao (evento_id, academia_id, aderiu, valor_taxa) 
                    VALUES (%s, %s, %s, %s)
                """, (evento_id, academia_id, aderiu, valor_taxa))
        else:
            # Se não tem valor_taxa ou está desaderindo, apenas atualizar aderiu
            cur.execute("""
                UPDATE eventos_competicoes_adesao SET aderiu = %s WHERE evento_id = %s AND academia_id = %s
            """, (aderiu, evento_id, academia_id))
            if cur.rowcount == 0:
                if tem_coluna_valor_taxa and valor_taxa is not None:
                    cur.execute("INSERT INTO eventos_competicoes_adesao (evento_id, academia_id, aderiu, valor_taxa) VALUES (%s, %s, %s, %s)",
                        (evento_id, academia_id, aderiu, valor_taxa))
                else:
                    cur.execute("INSERT INTO eventos_competicoes_adesao (evento_id, academia_id, aderiu) VALUES (%s, %s, %s)",
                        (evento_id, academia_id, aderiu))
        
        conn.commit()
        flash("Adesão atualizada." if aderiu else "Adesão removida.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.lista", academia_id=academia_id))


@bp_eventos_competicoes.route("/<int:evento_id>/inscritos")
@login_required
def inscritos(evento_id):
    """Gestor academia: lista de inscritos, incluir avulso, enviar à associação."""
    academia_id = request.args.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.lista"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se colunas de taxa existem
        tem_coluna_taxa, tem_coluna_valor_taxa = _colunas_taxa_existem(cur)
        if tem_coluna_taxa and tem_coluna_valor_taxa:
            cur.execute("""
                SELECT ec.id, ec.nome, ec.data_fim, ec.id_formulario, COALESCE(ec.tem_taxa, 0) as tem_taxa, 
                       ec.valor_taxa_sugerido, ea.valor_taxa
                FROM eventos_competicoes ec
                INNER JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s AND ea.aderiu = 1
                WHERE ec.id = %s
            """, (academia_id, evento_id))
        else:
            cur.execute("""
                SELECT ec.id, ec.nome, ec.data_fim, ec.id_formulario, 0 as tem_taxa, NULL as valor_taxa_sugerido, NULL as valor_taxa
                FROM eventos_competicoes ec
                INNER JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s AND ea.aderiu = 1
                WHERE ec.id = %s
            """, (academia_id, evento_id))
        ev = cur.fetchone()
        if not ev:
            flash("Evento não encontrado ou academia não aderiu.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))
        encerrado = _evento_encerrado(ev)
        # Garantir que tem_taxa seja int para comparação no template
        if 'tem_taxa' in ev:
            ev['tem_taxa'] = int(ev['tem_taxa']) if ev['tem_taxa'] else 0

        # Verificar se tabela de pagamentos existe
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
        
        if tabela_pagamentos_existe:
            cur.execute("""
                SELECT i.id, i.aluno_id, i.dados_form, i.inclusao_avulsa, i.status, i.created_at, i.data_envio,
                       a.nome as aluno_nome, a.foto as aluno_foto,
                       p.id as pagamento_id, p.valor as valor_taxa, 
                       COALESCE(p.pago_academia, 0) as pago_academia, 
                       COALESCE(p.pago_associacao, 0) as pago_associacao
                FROM eventos_competicoes_inscricoes i
                LEFT JOIN alunos a ON a.id = i.aluno_id
                LEFT JOIN eventos_competicoes_inscricoes_pagamentos p ON p.inscricao_id = i.id
                WHERE i.evento_id = %s AND i.academia_id = %s
                ORDER BY i.inclusao_avulsa ASC, a.nome ASC
            """, (evento_id, academia_id))
        else:
            cur.execute("""
                SELECT i.id, i.aluno_id, i.dados_form, i.inclusao_avulsa, i.status, i.created_at, i.data_envio,
                       a.nome as aluno_nome, a.foto as aluno_foto,
                       NULL as pagamento_id, NULL as valor_taxa, 0 as pago_academia, 0 as pago_associacao
                FROM eventos_competicoes_inscricoes i
                LEFT JOIN alunos a ON a.id = i.aluno_id
                WHERE i.evento_id = %s AND i.academia_id = %s
                ORDER BY i.inclusao_avulsa ASC, a.nome ASC
            """, (evento_id, academia_id))
        inscricoes = cur.fetchall()
        for i in inscricoes:
            if i.get("dados_form") and isinstance(i["dados_form"], str):
                try:
                    i["dados_form"] = json.loads(i["dados_form"])
                except Exception:
                    i["dados_form"] = {}
            # Garantir que pago_academia seja sempre int (0 ou 1)
            if "pago_academia" in i:
                i["pago_academia"] = int(i["pago_academia"]) if i["pago_academia"] else 0
            else:
                i["pago_academia"] = 0

        cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
            (ev["id_formulario"] or 0,))
        campos_form = cur.fetchall()

        # Calcular valores financeiros
        valor_associacao = float(ev.get("valor_taxa_sugerido") or 0)
        valor_academia = float(ev.get("valor_taxa") or 0)
        total_inscritos = len(inscricoes)
        
        # Calcular valores de pagamentos
        total_a_pagar = 0  # Valor que academia deve pagar à associação
        total_a_receber = 0  # Valor que academia recebe dos alunos
        total_pago = 0  # Valor já pago pelos alunos
        total_pendente = 0  # Valor pendente dos alunos
        
        if tabela_pagamentos_existe and ev["tem_taxa"] == 1:
            for insc in inscricoes:
                if insc.get("pagamento_id"):
                    valor_pagamento = float(insc.get("valor_taxa") or 0)
                    if insc.get("pago_academia"):
                        total_pago += valor_pagamento
                    else:
                        total_pendente += valor_pagamento
                    total_a_receber += valor_pagamento
        
        # Valor a pagar = valor da associação x quantidade de inscritos
        total_a_pagar = valor_associacao * total_inscritos
        
        # Receita = valor recebido confirmado - valor proporcional a pagar à associação
        # Contar quantos alunos têm pagamento confirmado
        total_inscritos_com_pagamento_confirmado = sum(1 for insc in inscricoes if insc.get("pago_academia"))
        valor_proporcional_a_pagar = valor_associacao * total_inscritos_com_pagamento_confirmado if total_inscritos_com_pagamento_confirmado > 0 else 0
        
        # Receita só aparece quando há pagamentos confirmados
        receita = total_pago - valor_proporcional_a_pagar if total_pago > 0 else 0
        
        # Verificar se há inscrições enviadas
        tem_inscricoes_enviadas = any(i.get("status") == "enviada" for i in inscricoes)
        
        # Verificar se usuário é associação (para mostrar botão cancelar envio)
        is_associacao = session.get("modo_painel") == "associacao" and (current_user.has_role("gestor_associacao") or current_user.has_role("admin"))

        return render_template("eventos_competicoes/inscritos.html",
            evento=ev, inscricoes=inscricoes, academia_id=academia_id, encerrado=encerrado, campos_form=campos_form,
            valor_associacao=valor_associacao, valor_academia=valor_academia,
            total_inscritos=total_inscritos, total_a_pagar=total_a_pagar, total_a_receber=total_a_receber,
            total_pago=total_pago, total_pendente=total_pendente, receita=receita,
            tabela_pagamentos_existe=tabela_pagamentos_existe,
            tem_inscricoes_enviadas=tem_inscricoes_enviadas, is_associacao=is_associacao,
            back_url=url_for("eventos_competicoes.lista", academia_id=academia_id))
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/<int:evento_id>/enviar-inscricoes", methods=["POST"])
@login_required
def enviar_inscricoes(evento_id):
    """Gestor academia marca inscrições como enviadas à associação. Valida que todas estão pagas se evento tem taxa."""
    academia_id = request.form.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.lista"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se colunas de taxa existem
        tem_coluna_taxa, tem_coluna_valor_taxa = _colunas_taxa_existem(cur)
        if tem_coluna_taxa and tem_coluna_valor_taxa:
            cur.execute("""
                SELECT ec.id, ec.data_fim, ec.tem_taxa, ec.valor_taxa_sugerido, ea.valor_taxa
                FROM eventos_competicoes ec
                LEFT JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s
                WHERE ec.id = %s
            """, (academia_id, evento_id))
        else:
            cur.execute("""
                SELECT ec.id, ec.data_fim, 0 as tem_taxa, NULL as valor_taxa_sugerido, NULL as valor_taxa
                FROM eventos_competicoes ec
                LEFT JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s
                WHERE ec.id = %s
            """, (academia_id, evento_id))
        ev = cur.fetchone()
        if not ev or _evento_encerrado(ev):
            flash("Evento encerrado.", "warning")
            return redirect(url_for("eventos_competicoes.lista"))

        # Se evento tem taxa, validar que todas as inscrições estão pagas
        tem_taxa_int = int(ev.get("tem_taxa") or 0) if ev.get("tem_taxa") else 0
        if tem_taxa_int == 1 and ev.get("valor_taxa"):
            # Verificar se tabela de pagamentos existe
            try:
                cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
                tabela_pagamentos_existe = cur.fetchone() is not None
            except Exception:
                tabela_pagamentos_existe = False
            
            if tabela_pagamentos_existe:
                cur.execute("""
                    SELECT COUNT(*) as total_inscricoes,
                           COUNT(CASE WHEN p.pago_academia = 1 THEN 1 END) as total_pagas
                    FROM eventos_competicoes_inscricoes i
                    LEFT JOIN eventos_competicoes_inscricoes_pagamentos p ON p.inscricao_id = i.id
                    WHERE i.evento_id = %s AND i.academia_id = %s AND i.status != 'enviada'
                """, (evento_id, academia_id))
                stats = cur.fetchone()
                total_inscricoes = stats.get("total_inscricoes", 0) or 0
                total_pagas = stats.get("total_pagas", 0) or 0
                
                if total_inscricoes > 0 and total_pagas < total_inscricoes:
                    pendentes = total_inscricoes - total_pagas
                    flash(f"Não é possível enviar. {pendentes} inscrição(ões) ainda não foram marcadas como pagas.", "danger")
                    return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # Enviar inscrições
        cur.execute("""
            UPDATE eventos_competicoes_inscricoes SET status = 'enviada', data_envio = NOW()
            WHERE evento_id = %s AND academia_id = %s AND status != 'enviada'
        """, (evento_id, academia_id))
        
        # Se evento tem taxa, criar ou atualizar registro de pagamento da academia para associação
        # Usar valor da associação (valor_taxa_sugerido) para calcular o que a academia deve pagar
        if tem_taxa_int == 1 and ev.get("valor_taxa_sugerido"):
            # Verificar se tabela de pagamentos existe
            try:
                cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
                tabela_pagamentos_academia_existe = cur.fetchone() is not None
            except Exception:
                tabela_pagamentos_academia_existe = False
            
            if tabela_pagamentos_academia_existe:
                # Contar quantas inscrições foram enviadas
                cur.execute("""
                    SELECT COUNT(*) as total_enviadas
                    FROM eventos_competicoes_inscricoes
                    WHERE evento_id = %s AND academia_id = %s AND status = 'enviada'
                """, (evento_id, academia_id))
                total_enviadas = cur.fetchone().get("total_enviadas", 0) or 0
                valor_total_esperado = float(ev.get("valor_taxa_sugerido", 0) or 0) * total_enviadas
                
                # Verificar se já existe registro de pagamento
                cur.execute("""
                    SELECT id, valor_pago FROM eventos_competicoes_academia_pagamentos
                    WHERE evento_id = %s AND academia_id = %s
                """, (evento_id, academia_id))
                pagamento_existente = cur.fetchone()
                
                if pagamento_existente:
                    # Atualizar valor total esperado
                    valor_pago_atual = float(pagamento_existente.get("valor_pago", 0) or 0)
                    valor_pendente = valor_total_esperado - valor_pago_atual
                    status_pagamento = "quitado" if valor_pendente <= 0 else ("parcial" if valor_pago_atual > 0 else "pendente")
                    cur.execute("""
                        UPDATE eventos_competicoes_academia_pagamentos
                        SET valor_total_esperado = %s, valor_pendente = %s, status = %s
                        WHERE id = %s
                    """, (valor_total_esperado, valor_pendente, status_pagamento, pagamento_existente["id"]))
                else:
                    # Criar novo registro
                    cur.execute("""
                        INSERT INTO eventos_competicoes_academia_pagamentos
                        (evento_id, academia_id, valor_total_esperado, valor_pendente, status)
                        VALUES (%s, %s, %s, %s, 'pendente')
                    """, (evento_id, academia_id, valor_total_esperado, valor_total_esperado))
        
        conn.commit()
        flash("Inscrições enviadas à associação.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))


@bp_eventos_competicoes.route("/<int:evento_id>/editar-inscricao/<int:inscricao_id>", methods=["GET", "POST"])
@login_required
def editar_inscricao(evento_id, inscricao_id):
    """Gestor academia edita dados de inscrição antes de enviar."""
    academia_id = request.args.get("academia_id", type=int) or request.form.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.lista"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar inscricao
        cur.execute("""
            SELECT i.id, i.aluno_id, i.dados_form, i.status, a.nome as aluno_nome
            FROM eventos_competicoes_inscricoes i
            LEFT JOIN alunos a ON a.id = i.aluno_id
            WHERE i.id = %s AND i.evento_id = %s AND i.academia_id = %s
        """, (inscricao_id, evento_id, academia_id))
        inscricao = cur.fetchone()
        if not inscricao:
            flash("Inscrição não encontrada.", "danger")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # Verificar permissão: gestor_academia, gestor_associacao ou admin podem editar
        pode_editar = (
            current_user.has_role("gestor_academia") or
            current_user.has_role("gestor_associacao") or
            current_user.has_role("admin")
        )
        if not pode_editar:
            flash("Você não tem permissão para editar inscrições.", "danger")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        if inscricao["status"] == "enviada":
            flash("Não é possível editar inscrição já enviada.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # Buscar evento e campos do formulário
        cur.execute("SELECT id, nome, id_formulario, data_fim FROM eventos_competicoes WHERE id = %s", (evento_id,))
        ev = cur.fetchone()
        if not ev or _evento_encerrado(ev):
            flash("Evento encerrado.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
            (ev["id_formulario"] or 0,))
        campos_form = cur.fetchall()
        
        # Verificar se o formulário tem peso e data_nascimento mas não tem categoria
        # Se sim, adicionar categoria automaticamente
        campos_chaves = [c["campo_chave"] for c in campos_form]
        tem_peso = "peso" in campos_chaves
        tem_data_nasc = "data_nascimento" in campos_chaves
        tem_categoria = "categoria" in campos_chaves
        
        if tem_peso and tem_data_nasc and not tem_categoria:
            try:
                # Adicionar categoria após peso
                ordem_peso = next((c["ordem"] for c in campos_form if c["campo_chave"] == "peso"), len(campos_form))
                from utils.formularios_campos import get_label
                label_categoria = get_label("categoria")
                cur.execute("""
                    INSERT INTO formularios_campos (formulario_id, campo_chave, label, ordem)
                    VALUES (%s, %s, %s, %s)
                """, (ev["id_formulario"], "categoria", label_categoria, ordem_peso + 1))
                conn.commit()
                # Recarregar campos
                cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
                    (ev["id_formulario"] or 0,))
                campos_form = cur.fetchall()
            except Exception:
                # Se já existe ou erro, continuar normalmente
                pass

        cur.execute("SELECT id, faixa, graduacao, categoria FROM graduacao ORDER BY id")
        graduacoes = cur.fetchall()
        cur.execute("SELECT TurmaID, Nome, Classificacao, DiasHorario FROM turmas WHERE id_academia = %s ORDER BY Nome", (academia_id,))
        turmas = cur.fetchall()
        cur.execute("SELECT id, nome FROM professores WHERE id_academia = %s AND ativo = 1 ORDER BY nome", (academia_id,))
        professores = cur.fetchall()
        cur.execute("SELECT id, nome FROM professores WHERE id_academia = %s AND ativo = 1 ORDER BY nome", (academia_id,))
        professores = cur.fetchall()
        cur.execute("SELECT id, nome FROM professores WHERE id_academia = %s AND ativo = 1 ORDER BY nome", (academia_id,))
        professores = cur.fetchall()
        
        # Buscar aluno para obter dados necessários para categorias
        aluno_edit = None
        if inscricao.get("aluno_id"):
            cur.execute("SELECT * FROM alunos WHERE id = %s", (inscricao["aluno_id"],))
            aluno_edit = cur.fetchone()
        
        # Buscar categorias disponíveis
        categorias_disponiveis = []
        if aluno_edit:
            peso = request.form.get("campo_peso") or aluno_edit.get("peso")
            data_nascimento = request.form.get("campo_data_nascimento") or aluno_edit.get("data_nascimento")
            genero = request.form.get("campo_sexo") or aluno_edit.get("sexo")
            
            if peso and data_nascimento and genero:
                try:
                    from datetime import datetime as dt
                    if isinstance(data_nascimento, str):
                        nasc = dt.strptime(data_nascimento[:10], "%Y-%m-%d").date()
                    else:
                        nasc = data_nascimento
                    hoje = date.today()
                    idade_ano_civil = hoje.year - nasc.year
                    genero_upper = (genero or "").upper()
                    peso_float = float(peso)
                    
                    if genero_upper in ("M", "F") and peso_float > 0:
                        # Buscar id_classe também se existir na tabela
                        try:
                            cur.execute("SHOW COLUMNS FROM categorias LIKE 'id_classe'")
                            tem_id_classe = cur.fetchone() is not None
                        except Exception:
                            tem_id_classe = False
                        
                        if tem_id_classe:
                            cur.execute("""
                                SELECT id, categoria, nome_categoria, id_classe, peso_min, peso_max, idade_min, idade_max
                                FROM categorias
                                WHERE genero = %s
                                AND (idade_min IS NULL OR %s >= idade_min)
                                AND (idade_max IS NULL OR %s <= idade_max)
                                AND (peso_min IS NULL OR %s >= peso_min)
                                AND (peso_max IS NULL OR %s <= peso_max)
                                ORDER BY nome_categoria
                            """, (genero_upper, idade_ano_civil, idade_ano_civil, peso_float, peso_float))
                        else:
                            cur.execute("""
                                SELECT id, categoria, nome_categoria, NULL as id_classe, peso_min, peso_max, idade_min, idade_max
                                FROM categorias
                                WHERE genero = %s
                                AND (idade_min IS NULL OR %s >= idade_min)
                                AND (idade_max IS NULL OR %s <= idade_max)
                                AND (peso_min IS NULL OR %s >= peso_min)
                                AND (peso_max IS NULL OR %s <= peso_max)
                                ORDER BY nome_categoria
                            """, (genero_upper, idade_ano_civil, idade_ano_civil, peso_float, peso_float))
                        categorias_disponiveis = cur.fetchall()
                except Exception:
                    categorias_disponiveis = []

        if request.method == "POST":
            dados = {}
            for c in campos_form:
                chave = c["campo_chave"]
                val = request.form.get(f"campo_{chave}", "")
                if isinstance(val, str):
                    val = val.strip()
                # Converter data de formato BR para ISO se necessário
                if chave in ("data_nascimento", "ultimo_exame_faixa", "rg_data_emissao", "data_cadastro_zempo") and val:
                    try:
                        from datetime import datetime as dt
                        val_str = val[:10].strip()
                        if val_str and '/' in val_str:
                            partes = val_str.split('/')
                            if len(partes) == 3:
                                val = f"{partes[2]}-{partes[1]}-{partes[0]}"
                    except Exception:
                        pass
                dados[chave] = val
            cur.execute("""
                UPDATE eventos_competicoes_inscricoes SET dados_form = %s WHERE id = %s
            """, (json.dumps(dados, ensure_ascii=False), inscricao_id))
            conn.commit()
            flash("Inscrição atualizada.", "success")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # GET: montar valores iniciais
        valores_iniciais = {}
        dados_atuais = {}
        if inscricao.get("dados_form"):
            if isinstance(inscricao["dados_form"], str):
                try:
                    dados_atuais = json.loads(inscricao["dados_form"])
                except Exception:
                    pass
            else:
                dados_atuais = inscricao["dados_form"]
        for c in campos_form:
            chave = c["campo_chave"]
            v = dados_atuais.get(chave, "")
            # Converter datas para formato BR
            if chave in ("data_nascimento", "ultimo_exame_faixa", "rg_data_emissao", "data_cadastro_zempo") and v:
                try:
                    from datetime import datetime as dt
                    # Tentar converter de ISO para BR
                    if isinstance(v, str) and '-' in v:
                        d = dt.strptime(v[:10], "%Y-%m-%d")
                        v = d.strftime("%d/%m/%Y")
                except Exception:
                    pass
            elif chave == "id_academia":
                # Usar academia_id atual
                v = academia_id or v
            valores_iniciais[chave] = v

        # Buscar nome da academia para exibição
        academia_nome = None
        if academia_id:
            cur.execute("SELECT nome FROM academias WHERE id = %s", (academia_id,))
            ac_row = cur.fetchone()
            academia_nome = ac_row["nome"] if ac_row else None

        return render_template("eventos_competicoes/editar_inscricao.html",
            evento=ev, inscricao=inscricao, academia_id=academia_id, academia_nome=academia_nome,
            campos_form=campos_form, valores_iniciais=valores_iniciais, graduacoes=graduacoes, turmas=turmas,
            professores=professores, categorias_disponiveis=categorias_disponiveis, aluno=aluno_edit,
            back_url=url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/<int:evento_id>/marcar-pagamento/<int:inscricao_id>", methods=["POST"])
@login_required
def marcar_pagamento(evento_id, inscricao_id):
    """Academia marca pagamento de uma inscrição como pago."""
    academia_id = request.form.get("academia_id", type=int) or request.args.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.lista", academia_id=academia_id))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se tabela de pagamentos existe
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
        
        if not tabela_pagamentos_existe:
            flash("Sistema de pagamentos não está disponível.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        # Verificar se a inscrição pertence à academia
        cur.execute("""
            SELECT i.id, i.status, p.id as pagamento_id
            FROM eventos_competicoes_inscricoes i
            LEFT JOIN eventos_competicoes_inscricoes_pagamentos p ON p.inscricao_id = i.id
            WHERE i.id = %s AND i.academia_id = %s AND i.evento_id = %s
        """, (inscricao_id, academia_id, evento_id))
        inscricao = cur.fetchone()
        
        if not inscricao:
            flash("Inscrição não encontrada.", "danger")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        if inscricao.get("status") == "enviada":
            flash("Não é possível alterar pagamento de inscrição já enviada.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        if not inscricao.get("pagamento_id"):
            flash("Esta inscrição não possui registro de pagamento.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        # Buscar informações do evento para obter valor_taxa_sugerido
        try:
            cur.execute("SHOW COLUMNS FROM eventos_competicoes LIKE 'tem_taxa'")
            tem_coluna_taxa = cur.fetchone() is not None
        except Exception:
            tem_coluna_taxa = False
        
        valor_taxa_associacao = 0.0
        if tem_coluna_taxa:
            cur.execute("""
                SELECT tem_taxa, valor_taxa_sugerido
                FROM eventos_competicoes
                WHERE id = %s
            """, (evento_id,))
            evento_info = cur.fetchone()
            if evento_info and evento_info.get("tem_taxa"):
                valor_taxa_associacao = float(evento_info.get("valor_taxa_sugerido", 0) or 0)
        
        # Verificar se já estava marcado como pago antes (para não somar duas vezes)
        cur.execute("""
            SELECT pago_academia FROM eventos_competicoes_inscricoes_pagamentos
            WHERE id = %s
        """, (inscricao["pagamento_id"],))
        pagamento_atual = cur.fetchone()
        ja_estava_pago = pagamento_atual and pagamento_atual.get("pago_academia") == 1
        
        # Marcar como pago na tabela de pagamentos de inscrições
        cur.execute("""
            UPDATE eventos_competicoes_inscricoes_pagamentos
            SET pago_academia = 1, data_pagamento_academia = NOW()
            WHERE id = %s
        """, (inscricao["pagamento_id"],))
        
        # Atualizar tabela de pagamentos das academias (se evento tem taxa e valor > 0)
        if valor_taxa_associacao > 0 and not ja_estava_pago:
            # Verificar se tabela existe
            try:
                cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
                tabela_pagamentos_academias_existe = cur.fetchone() is not None
            except Exception:
                tabela_pagamentos_academias_existe = False
            
            if tabela_pagamentos_academias_existe:
                # Buscar registro existente ou criar novo
                cur.execute("""
                    SELECT id, valor_pago, valor_total_esperado, valor_pendente, status
                    FROM eventos_competicoes_academia_pagamentos
                    WHERE evento_id = %s AND academia_id = %s
                """, (evento_id, academia_id))
                pagamento_academia = cur.fetchone()
                
                if pagamento_academia:
                    # Atualizar valor_pago somando o valor da taxa da associação
                    novo_valor_pago = float(pagamento_academia.get("valor_pago", 0) or 0) + valor_taxa_associacao
                    valor_total_esperado = float(pagamento_academia.get("valor_total_esperado", 0) or 0)
                    novo_valor_pendente = valor_total_esperado - novo_valor_pago
                    
                    # Atualizar status
                    if novo_valor_pendente <= 0:
                        novo_status = "quitado"
                    elif novo_valor_pago > 0:
                        novo_status = "parcial"
                    else:
                        novo_status = "pendente"
                    
                    cur.execute("""
                        UPDATE eventos_competicoes_academia_pagamentos
                        SET valor_pago = %s, valor_pendente = %s, status = %s
                        WHERE id = %s
                    """, (novo_valor_pago, novo_valor_pendente, novo_status, pagamento_academia["id"]))
                else:
                    # Criar novo registro se não existir
                    # Contar total de inscrições enviadas para calcular valor_total_esperado
                    cur.execute("""
                        SELECT COUNT(*) as total
                        FROM eventos_competicoes_inscricoes
                        WHERE evento_id = %s AND academia_id = %s AND status = 'enviada'
                    """, (evento_id, academia_id))
                    total_result = cur.fetchone()
                    total_inscricoes = (total_result.get("total", 0) or 0) if total_result else 0
                    
                    # Se não há inscrições enviadas ainda, contar as confirmadas
                    if total_inscricoes == 0:
                        cur.execute("""
                            SELECT COUNT(*) as total
                            FROM eventos_competicoes_inscricoes
                            WHERE evento_id = %s AND academia_id = %s AND status = 'confirmada'
                        """, (evento_id, academia_id))
                        total_result = cur.fetchone()
                        total_inscricoes = (total_result.get("total", 0) or 0) if total_result else 0
                    
                    valor_total_esperado = valor_taxa_associacao * total_inscricoes
                    novo_valor_pago = valor_taxa_associacao  # Primeiro pagamento confirmado
                    novo_valor_pendente = valor_total_esperado - novo_valor_pago
                    
                    if novo_valor_pendente <= 0:
                        novo_status = "quitado"
                    else:
                        novo_status = "parcial"
                    
                    cur.execute("""
                        INSERT INTO eventos_competicoes_academia_pagamentos
                        (evento_id, academia_id, valor_total_esperado, valor_pago, valor_pendente, status)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (evento_id, academia_id, valor_total_esperado, novo_valor_pago, novo_valor_pendente, novo_status))
        
        conn.commit()
        flash("Pagamento confirmado com sucesso.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))


@bp_eventos_competicoes.route("/<int:evento_id>/criar-pagamento/<int:inscricao_id>", methods=["POST"])
@login_required
def criar_pagamento(evento_id, inscricao_id):
    """Academia cria registro de pagamento para uma inscrição que não tem."""
    academia_id = request.form.get("academia_id", type=int) or request.args.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se tabela de pagamentos existe, se não, criar
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
        
        if not tabela_pagamentos_existe:
            # Tentar criar a tabela
            try:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS eventos_competicoes_inscricoes_pagamentos (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        inscricao_id INT NOT NULL,
                        valor DECIMAL(10,2) NOT NULL,
                        pago_academia TINYINT(1) NOT NULL DEFAULT 0,
                        pago_associacao TINYINT(1) NOT NULL DEFAULT 0,
                        data_pagamento_academia DATETIME NULL DEFAULT NULL,
                        data_pagamento_associacao DATETIME NULL DEFAULT NULL,
                        observacoes TEXT NULL DEFAULT NULL,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        INDEX idx_pag_inscricao (inscricao_id)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_uca1400_ai_ci
                """)
                conn.commit()
                tabela_pagamentos_existe = True
            except Exception as e:
                try:
                    current_app.logger.error(f"Erro ao criar tabela de pagamentos: {e}")
                except:
                    pass
        
        if not tabela_pagamentos_existe:
            flash("Sistema de pagamentos não está disponível. Execute a migration primeiro.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        # Verificar se a inscrição pertence à academia
        cur.execute("""
            SELECT i.id, i.status
            FROM eventos_competicoes_inscricoes i
            WHERE i.id = %s AND i.academia_id = %s AND i.evento_id = %s
        """, (inscricao_id, academia_id, evento_id))
        inscricao = cur.fetchone()
        
        if not inscricao:
            flash("Inscrição não encontrada.", "danger")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        if inscricao.get("status") == "enviada":
            flash("Não é possível criar pagamento para inscrição já enviada.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        # Verificar se já existe registro de pagamento
        cur.execute("SELECT id FROM eventos_competicoes_inscricoes_pagamentos WHERE inscricao_id = %s", (inscricao_id,))
        if cur.fetchone():
            flash("Registro de pagamento já existe.", "info")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        # Buscar valor da taxa do evento
        tem_coluna_taxa, tem_coluna_valor_taxa = _colunas_taxa_existem(cur)
        if not tem_coluna_taxa or not tem_coluna_valor_taxa:
            flash("Sistema de taxas não está configurado.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        cur.execute("""
            SELECT COALESCE(ec.tem_taxa, 0) as tem_taxa, ea.valor_taxa
            FROM eventos_competicoes ec
            LEFT JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s
            WHERE ec.id = %s
        """, (academia_id, evento_id))
        ev_info = cur.fetchone()
        
        if not ev_info or not ev_info.get("tem_taxa") or not ev_info.get("valor_taxa"):
            flash("Evento não possui taxa configurada ou valor não definido.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        valor = float(ev_info.get("valor_taxa"))
        
        # Criar registro de pagamento
        cur.execute("""
            INSERT INTO eventos_competicoes_inscricoes_pagamentos (inscricao_id, valor)
            VALUES (%s, %s)
        """, (inscricao_id, valor))
        conn.commit()
        flash(f"Registro de pagamento criado (R$ {valor:.2f}). Agora você pode confirmar o pagamento.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))


@bp_eventos_competicoes.route("/<int:evento_id>/cancelar-pagamento/<int:inscricao_id>", methods=["POST"])
@login_required
def cancelar_pagamento(evento_id, inscricao_id):
    """Academia cancela confirmação de pagamento de uma inscrição."""
    academia_id = request.form.get("academia_id", type=int) or request.args.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se tabela de pagamentos existe
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
        
        if not tabela_pagamentos_existe:
            flash("Sistema de pagamentos não está disponível.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        # Verificar se a inscrição pertence à academia
        cur.execute("""
            SELECT i.id, i.status, p.id as pagamento_id, p.pago_academia
            FROM eventos_competicoes_inscricoes i
            LEFT JOIN eventos_competicoes_inscricoes_pagamentos p ON p.inscricao_id = i.id
            WHERE i.id = %s AND i.academia_id = %s AND i.evento_id = %s
        """, (inscricao_id, academia_id, evento_id))
        inscricao = cur.fetchone()
        
        if not inscricao:
            flash("Inscrição não encontrada.", "danger")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        if inscricao.get("status") == "enviada":
            flash("Não é possível alterar pagamento de inscrição já enviada.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        if not inscricao.get("pagamento_id"):
            flash("Registro de pagamento não encontrado.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        if not inscricao.get("pago_academia"):
            flash("Pagamento já está pendente.", "info")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
        
        # Buscar informações do evento para obter valor_taxa_sugerido
        try:
            cur.execute("SHOW COLUMNS FROM eventos_competicoes LIKE 'tem_taxa'")
            tem_coluna_taxa = cur.fetchone() is not None
        except Exception:
            tem_coluna_taxa = False
        
        valor_taxa_associacao = 0.0
        if tem_coluna_taxa:
            cur.execute("""
                SELECT tem_taxa, valor_taxa_sugerido
                FROM eventos_competicoes
                WHERE id = %s
            """, (evento_id,))
            evento_info = cur.fetchone()
            if evento_info and evento_info.get("tem_taxa"):
                valor_taxa_associacao = float(evento_info.get("valor_taxa_sugerido", 0) or 0)
        
        # Cancelar confirmação de pagamento na tabela de pagamentos de inscrições
        cur.execute("""
            UPDATE eventos_competicoes_inscricoes_pagamentos
            SET pago_academia = 0, data_pagamento_academia = NULL
            WHERE id = %s
        """, (inscricao["pagamento_id"],))
        
        # Atualizar tabela de pagamentos das academias (reverter valor_pago)
        if valor_taxa_associacao > 0:
            # Verificar se tabela existe
            try:
                cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
                tabela_pagamentos_academias_existe = cur.fetchone() is not None
            except Exception:
                tabela_pagamentos_academias_existe = False
            
            if tabela_pagamentos_academias_existe:
                # Buscar registro existente
                cur.execute("""
                    SELECT id, valor_pago, valor_total_esperado, valor_pendente, status
                    FROM eventos_competicoes_academia_pagamentos
                    WHERE evento_id = %s AND academia_id = %s
                """, (evento_id, academia_id))
                pagamento_academia = cur.fetchone()
                
                if pagamento_academia:
                    # Reverter valor_pago subtraindo o valor da taxa da associação
                    valor_pago_atual = float(pagamento_academia.get("valor_pago", 0) or 0)
                    novo_valor_pago = max(0, valor_pago_atual - valor_taxa_associacao)  # Não deixar negativo
                    valor_total_esperado = float(pagamento_academia.get("valor_total_esperado", 0) or 0)
                    novo_valor_pendente = valor_total_esperado - novo_valor_pago
                    
                    # Atualizar status
                    if novo_valor_pendente <= 0:
                        novo_status = "quitado"
                    elif novo_valor_pago > 0:
                        novo_status = "parcial"
                    else:
                        novo_status = "pendente"
                    
                    cur.execute("""
                        UPDATE eventos_competicoes_academia_pagamentos
                        SET valor_pago = %s, valor_pendente = %s, status = %s
                        WHERE id = %s
                    """, (novo_valor_pago, novo_valor_pendente, novo_status, pagamento_academia["id"]))
        
        conn.commit()
        flash("Confirmação de pagamento cancelada. O pagamento voltou ao status pendente.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))


@bp_eventos_competicoes.route("/<int:evento_id>/cancelar-inscricao/<int:inscricao_id>", methods=["POST"])
@login_required
def cancelar_inscricao(evento_id, inscricao_id):
    """Gestor academia cancela inscrição antes de enviar."""
    academia_id = request.form.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.lista"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar inscricao
        cur.execute("""
            SELECT i.id, i.status FROM eventos_competicoes_inscricoes i
            WHERE i.id = %s AND i.evento_id = %s AND i.academia_id = %s
        """, (inscricao_id, evento_id, academia_id))
        inscricao = cur.fetchone()
        if not inscricao:
            flash("Inscrição não encontrada.", "danger")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        if inscricao["status"] == "enviada":
            flash("Não é possível cancelar inscrição já enviada.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # Verificar se há pagamento pago
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
            if tabela_pagamentos_existe:
                cur.execute("""
                    SELECT id, pago_academia FROM eventos_competicoes_inscricoes_pagamentos 
                    WHERE inscricao_id = %s
                """, (inscricao_id,))
                pagamento = cur.fetchone()
                if pagamento and pagamento.get("pago_academia"):
                    flash("Não é possível cancelar inscrição com pagamento confirmado. Cancele o pagamento primeiro.", "warning")
                    return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))
                # Remover pagamento associado se existir (só se não estiver pago)
                if pagamento:
                    cur.execute("DELETE FROM eventos_competicoes_inscricoes_pagamentos WHERE inscricao_id = %s", (inscricao_id,))
        except Exception:
            pass
        
        cur.execute("DELETE FROM eventos_competicoes_inscricoes WHERE id = %s", (inscricao_id,))
        conn.commit()
        flash("Inscrição cancelada.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))


@bp_eventos_competicoes.route("/<int:evento_id>/incluir-avulso", methods=["GET", "POST"])
@login_required
def incluir_avulso(evento_id):
    """Gestor academia inclui aluno avulso (manual) com formulário completo."""
    academia_id = request.args.get("academia_id", type=int) or request.form.get("academia_id", type=int)
    aluno_id = request.args.get("aluno_id", type=int) or request.form.get("aluno_id", type=int)
    
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.lista"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT ec.id, ec.nome, ec.data_fim, ec.id_formulario
            FROM eventos_competicoes ec
            INNER JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s AND ea.aderiu = 1
            WHERE ec.id = %s
        """, (academia_id, evento_id))
        ev = cur.fetchone()
        if not ev or _evento_encerrado(ev):
            flash("Evento não encontrado ou encerrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))

        cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
            (ev["id_formulario"] or 0,))
        campos_form = cur.fetchall()
        if not campos_form:
            flash("Formulário sem campos configurados.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # Buscar nome da academia
        cur.execute("SELECT nome FROM academias WHERE id = %s", (academia_id,))
        ac_row = cur.fetchone()
        academia_nome = ac_row.get("nome") if ac_row else "Academia"

        # Se não tem aluno_id, mostrar seleção de alunos
        if not aluno_id:
            cur.execute("SELECT id, nome FROM alunos WHERE id_academia = %s ORDER BY nome", (academia_id,))
            alunos = cur.fetchall()
            return render_template("eventos_competicoes/incluir_avulso.html",
                evento=ev, alunos=alunos, academia_id=academia_id,
                back_url=url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # Buscar dados do aluno
        cur.execute("SELECT * FROM alunos WHERE id = %s AND id_academia = %s", (aluno_id, academia_id))
        aluno = cur.fetchone()
        if not aluno:
            flash("Aluno não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.incluir_avulso", evento_id=evento_id, academia_id=academia_id))

        # Verificar se já está inscrito
        cur.execute("SELECT 1 FROM eventos_competicoes_inscricoes WHERE evento_id=%s AND academia_id=%s AND aluno_id=%s",
            (evento_id, academia_id, aluno_id))
        if cur.fetchone():
            flash("Aluno já inscrito neste evento.", "warning")
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # Buscar graduações e turmas
        cur.execute("SELECT id, faixa, graduacao, categoria FROM graduacao ORDER BY id")
        graduacoes = cur.fetchall()
        cur.execute("SELECT TurmaID, Nome, Classificacao, DiasHorario FROM turmas WHERE id_academia = %s ORDER BY Nome", (academia_id,))
        turmas = cur.fetchall()

        # Verificar se formulário tem categoria e buscar categorias disponíveis
        campos_chaves = [c["campo_chave"] for c in campos_form]
        tem_categoria = "categoria" in campos_chaves
        categorias_disponiveis = []
        
        if tem_categoria and aluno.get("peso") and aluno.get("data_nascimento") and aluno.get("sexo"):
            try:
                from datetime import datetime as dt
                nasc = dt.strptime(str(aluno["data_nascimento"])[:10], "%Y-%m-%d").date()
                hoje = date.today()
                idade_ano_civil = hoje.year - nasc.year
                genero_upper = (aluno.get("sexo") or "").upper()
                peso_float = float(aluno.get("peso") or 0)
                
                if genero_upper in ("M", "F") and peso_float > 0:
                    genero_db = "MASCULINO" if genero_upper == "M" else "FEMININO" if genero_upper == "F" else genero_upper
                    try:
                        cur.execute("SHOW COLUMNS FROM categorias LIKE 'id_classe'")
                        tem_id_classe = cur.fetchone() is not None
                    except Exception:
                        tem_id_classe = False
                    
                    if tem_id_classe:
                        cur.execute("""
                            SELECT id, categoria, nome_categoria, id_classe, peso_min, peso_max, idade_min, idade_max
                            FROM categorias
                            WHERE UPPER(genero) = UPPER(%s)
                            AND (
                                (idade_min IS NULL OR %s >= idade_min)
                                AND (idade_max IS NULL OR %s <= idade_max)
                            )
                            AND (
                                (peso_min IS NULL OR %s >= peso_min)
                                AND (peso_max IS NULL OR %s <= peso_max)
                            )
                            ORDER BY nome_categoria
                        """, (genero_db, idade_ano_civil, idade_ano_civil, peso_float, peso_float))
                    else:
                        cur.execute("""
                            SELECT id, categoria, nome_categoria, NULL as id_classe, peso_min, peso_max, idade_min, idade_max
                            FROM categorias
                            WHERE UPPER(genero) = UPPER(%s)
                            AND (
                                (idade_min IS NULL OR %s >= idade_min)
                                AND (idade_max IS NULL OR %s <= idade_max)
                            )
                            AND (
                                (peso_min IS NULL OR %s >= peso_min)
                                AND (peso_max IS NULL OR %s <= peso_max)
                            )
                            ORDER BY nome_categoria
                        """, (genero_db, idade_ano_civil, idade_ano_civil, peso_float, peso_float))
                    categorias_disponiveis = cur.fetchall()
            except Exception:
                categorias_disponiveis = []

        # POST: Processar formulário completo
        if request.method == "POST" and request.form.get("formulario_completo"):
            # Processar categorias (pode ser múltipla)
            categorias_selecionadas = request.form.getlist("campo_categoria[]")
            
            # Coletar dados do formulário
            dados = {}
            for c in campos_form:
                chave = c["campo_chave"]
                if chave == "categoria":
                    continue  # Categoria será tratada separadamente
                val = request.form.get(f"campo_{chave}", "")
                if isinstance(val, str):
                    val = val.strip()
                dados[chave] = val

            # Se há categorias selecionadas, criar uma inscrição para cada uma
            if categorias_selecionadas:
                inscricoes_criadas = 0
                for categoria_nome in categorias_selecionadas:
                    dados_cat = dados.copy()
                    dados_cat["categoria"] = categoria_nome
                    cur.execute("""
                        INSERT INTO eventos_competicoes_inscricoes (evento_id, academia_id, aluno_id, usuario_inscricao_id, dados_form, inclusao_avulsa, status)
                        VALUES (%s, %s, %s, %s, %s, 1, 'confirmada')
                    """, (evento_id, academia_id, aluno_id, current_user.id, json.dumps(dados_cat, ensure_ascii=False)))
                    inscricao_id = cur.lastrowid
                    _criar_pagamento_inscricao(cur, inscricao_id, evento_id, academia_id)
                    inscricoes_criadas += 1
                conn.commit()
                if inscricoes_criadas > 1:
                    flash(f"{inscricoes_criadas} inscrições criadas com sucesso! Uma para cada categoria selecionada.", "success")
                else:
                    flash("Inscrição criada com sucesso!", "success")
            else:
                # Sem categoria, criar uma inscrição normal
                cur.execute("""
                    INSERT INTO eventos_competicoes_inscricoes (evento_id, academia_id, aluno_id, usuario_inscricao_id, dados_form, inclusao_avulsa, status)
                    VALUES (%s, %s, %s, %s, %s, 1, 'confirmada')
                """, (evento_id, academia_id, aluno_id, current_user.id, json.dumps(dados, ensure_ascii=False)))
                inscricao_id = cur.lastrowid
                _criar_pagamento_inscricao(cur, inscricao_id, evento_id, academia_id)
                conn.commit()
                flash("Inscrição criada com sucesso!", "success")
            
            return redirect(url_for("eventos_competicoes.inscritos", evento_id=evento_id, academia_id=academia_id))

        # GET: Montar valores iniciais do aluno (mapear col DB -> form)
        REV_MAP = {"rua": "endereco", "tel_celular": "telefone_celular", "tel_residencial": "telefone_residencial",
                   "tel_comercial": "telefone_comercial", "tel_outro": "telefone_outro", "telefone": "telefone_celular"}
        valores_iniciais = {}
        for c in campos_form:
            chave = c["campo_chave"]
            v = aluno.get(chave)
            if v is None and chave in ("endereco",):
                v = aluno.get("rua")
            elif v is None and chave == "telefone_celular":
                v = aluno.get("tel_celular") or aluno.get("telefone")
            elif v is None and chave in ("telefone_residencial", "telefone_comercial", "telefone_outro"):
                v = aluno.get("tel_" + chave.split("_")[1])
            if hasattr(v, "strftime"):
                # Para input type="text" com formato BR, converter para BR
                v = v.strftime("%d/%m/%Y") if v else ""
            elif chave == "id_academia":
                # Se vier ID, usar o ID da academia atual
                v = academia_id or v
            elif chave == "graduacao_id" and v:
                # Manter ID para o value (o select já mostra o nome)
                pass
            elif chave == "TurmaID" and v:
                # Manter ID para o value (o select já mostra o nome)
                pass
            valores_iniciais[chave] = v or ""

        return render_template("eventos_competicoes/incluir_avulso_form.html",
            evento=ev, aluno=aluno, academia_id=academia_id, academia_nome=academia_nome,
            campos_form=campos_form, valores_iniciais=valores_iniciais,
            graduacoes=graduacoes, turmas=turmas, professores=professores,
            categorias_disponiveis=categorias_disponiveis,
            back_url=url_for("eventos_competicoes.incluir_avulso", evento_id=evento_id, academia_id=academia_id))
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/<int:evento_id>/inscrever", methods=["GET", "POST"])
@login_required
def inscrever(evento_id):
    """Aluno ou responsável realiza inscrição preenchendo o formulário."""
    academia_id = request.args.get("academia_id", type=int) or request.form.get("academia_id", type=int)
    aluno_id = request.args.get("aluno_id", type=int) or request.form.get("aluno_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("painel.home"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT ec.id, ec.nome, ec.data_fim, ec.id_formulario
            FROM eventos_competicoes ec
            INNER JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s AND ea.aderiu = 1
            WHERE ec.id = %s
        """, (academia_id, evento_id))
        ev = cur.fetchone()
        if not ev:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("painel.home"))
        if _evento_encerrado(ev):
            flash("Inscrições encerradas.", "warning")
            return redirect(url_for("painel.home"))

        # Aluno ou responsável: verificar permissão
        alunos_permitidos = []
        if current_user.has_role("aluno"):
            cur.execute("SELECT id FROM alunos WHERE id = %s AND id_academia = %s", (aluno_id or 0, academia_id))
            if cur.fetchone():
                alunos_permitidos = [aluno_id]
        elif current_user.has_role("responsavel"):
            cur.execute("SELECT aluno_id FROM responsavel_alunos WHERE usuario_id = %s", (current_user.id,))
            ids_resp = [r["aluno_id"] for r in cur.fetchall()]
            if ids_resp:
                ph = ",".join(["%s"] * len(ids_resp))
                cur.execute(f"SELECT id FROM alunos WHERE id IN ({ph}) AND id_academia = %s", tuple(ids_resp) + (academia_id,))
                alunos_permitidos = [r["id"] for r in cur.fetchall()]
            else:
                alunos_permitidos = []
        else:
            flash("Acesso negado.", "danger")
            return redirect(url_for("painel.home"))

        if not alunos_permitidos:
            flash("Nenhum aluno vinculado para inscrição.", "danger")
            return redirect(url_for("painel.home"))

        # Se responsável tem vários alunos, precisa escolher
        if len(alunos_permitidos) > 1 and not aluno_id:
            cur.execute("SELECT id, nome FROM alunos WHERE id IN (%s) ORDER BY nome" % ",".join(["%s"] * len(alunos_permitidos)),
                tuple(alunos_permitidos))
            alunos_sel = cur.fetchall()
            return render_template("eventos_competicoes/escolher_aluno.html",
                evento=ev, alunos=alunos_sel, academia_id=academia_id, back_url=request.referrer or url_for("painel.home"))

        aluno_id = aluno_id or alunos_permitidos[0]

        cur.execute("SELECT * FROM alunos WHERE id = %s", (aluno_id,))
        aluno = cur.fetchone()
        if not aluno or aluno_id not in alunos_permitidos:
            flash("Aluno não encontrado.", "danger")
            return redirect(url_for("painel.home"))

        # Removida verificação de "aluno já inscrito" pois agora pode haver múltiplas inscrições (uma por categoria)

        cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
            (ev["id_formulario"] or 0,))
        campos_form = cur.fetchall()
        if not campos_form:
            flash("Formulário sem campos configurados.", "warning")
            return redirect(request.referrer or url_for("painel.home"))
        
        # Verificar se o formulário tem peso, data_nascimento e sexo mas não tem categoria
        # Se sim, adicionar categoria automaticamente
        campos_chaves = [c["campo_chave"] for c in campos_form]
        tem_peso = "peso" in campos_chaves
        tem_data_nasc = "data_nascimento" in campos_chaves
        tem_sexo = "sexo" in campos_chaves
        tem_categoria = "categoria" in campos_chaves
        
        if tem_peso and tem_data_nasc and tem_sexo and not tem_categoria:
            try:
                # Adicionar categoria após peso
                ordem_peso = next((c["ordem"] for c in campos_form if c["campo_chave"] == "peso"), len(campos_form))
                from utils.formularios_campos import get_label
                label_categoria = get_label("categoria")
                cur.execute("""
                    INSERT INTO formularios_campos (formulario_id, campo_chave, label, ordem)
                    VALUES (%s, %s, %s, %s)
                """, (ev["id_formulario"], "categoria", label_categoria, ordem_peso + 1))
                conn.commit()
                # Recarregar campos
                cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
                    (ev["id_formulario"] or 0,))
                campos_form = cur.fetchall()
                # Atualizar lista de chaves
                campos_chaves = [c["campo_chave"] for c in campos_form]
                tem_categoria = "categoria" in campos_chaves
            except Exception as e:
                # Se já existe ou erro, continuar normalmente
                import traceback
                print(f"Erro ao adicionar categoria ao formulário: {e}")
                print(traceback.format_exc())
                pass

        cur.execute("SELECT id, faixa, graduacao, categoria FROM graduacao ORDER BY id")
        graduacoes = cur.fetchall()
        cur.execute("SELECT TurmaID, Nome, Classificacao, DiasHorario FROM turmas WHERE id_academia = %s ORDER BY Nome", (academia_id,))
        turmas = cur.fetchall()
        
        # Buscar categorias disponíveis se aluno tem peso, data_nascimento e sexo
        categorias_disponiveis = []
        if aluno.get("peso") and aluno.get("data_nascimento") and aluno.get("sexo"):
            try:
                from datetime import datetime as dt
                nasc = dt.strptime(str(aluno["data_nascimento"])[:10], "%Y-%m-%d").date()
                hoje = date.today()
                idade_ano_civil = hoje.year - nasc.year
                genero_upper = (aluno.get("sexo") or "").upper()
                peso_float = float(aluno.get("peso") or 0)
                
                if genero_upper in ("M", "F") and peso_float > 0:
                    # Mapear M/F para MASCULINO/FEMININO
                    genero_db = "MASCULINO" if genero_upper == "M" else "FEMININO" if genero_upper == "F" else genero_upper
                    # Buscar id_classe também se existir na tabela
                    try:
                        cur.execute("SHOW COLUMNS FROM categorias LIKE 'id_classe'")
                        tem_id_classe = cur.fetchone() is not None
                    except Exception:
                        tem_id_classe = False
                    
                    if tem_id_classe:
                        cur.execute("""
                            SELECT id, categoria, nome_categoria, id_classe, peso_min, peso_max, idade_min, idade_max
                            FROM categorias
                            WHERE UPPER(genero) = UPPER(%s)
                            AND (
                                (idade_min IS NULL OR %s >= idade_min)
                                AND (idade_max IS NULL OR %s <= idade_max)
                            )
                            AND (
                                (peso_min IS NULL OR %s >= peso_min)
                                AND (peso_max IS NULL OR %s <= peso_max)
                            )
                            ORDER BY nome_categoria
                        """, (genero_db, idade_ano_civil, idade_ano_civil, peso_float, peso_float))
                    else:
                        cur.execute("""
                            SELECT id, categoria, nome_categoria, NULL as id_classe, peso_min, peso_max, idade_min, idade_max
                            FROM categorias
                            WHERE UPPER(genero) = UPPER(%s)
                            AND (
                                (idade_min IS NULL OR %s >= idade_min)
                                AND (idade_max IS NULL OR %s <= idade_max)
                            )
                            AND (
                                (peso_min IS NULL OR %s >= peso_min)
                                AND (peso_max IS NULL OR %s <= peso_max)
                            )
                            ORDER BY nome_categoria
                        """, (genero_db, idade_ano_civil, idade_ano_civil, peso_float, peso_float))
                    categorias_disponiveis = cur.fetchall()
            except Exception:
                categorias_disponiveis = []

        if request.method == "POST":
            # Processar categorias (pode ser múltipla)
            categorias_selecionadas = request.form.getlist("campo_categoria[]")
            
            # Coletar dados base do formulário
            dados_base = {}
            for c in campos_form:
                chave = c["campo_chave"]
                if chave == "categoria":
                    continue  # Categoria será tratada separadamente
                val = request.form.get(f"campo_{chave}", "")
                if isinstance(val, str):
                    val = val.strip()
                dados_base[chave] = val
            
            # Verificar inscrições existentes para prevenir duplicatas
            cur.execute("""
                SELECT id, dados_form, status
                FROM eventos_competicoes_inscricoes
                WHERE evento_id = %s AND academia_id = %s AND aluno_id = %s
            """, (evento_id, academia_id, aluno_id))
            inscricoes_existentes = cur.fetchall()
            
            # Extrair categorias já inscritas
            categorias_ja_inscritas = set()
            for insc_existente in inscricoes_existentes:
                if insc_existente.get("dados_form"):
                    try:
                        dados_existentes = json.loads(insc_existente["dados_form"]) if isinstance(insc_existente["dados_form"], str) else insc_existente["dados_form"]
                        if dados_existentes.get("categoria"):
                            categorias_ja_inscritas.add(dados_existentes["categoria"])
                    except Exception:
                        pass
            
            # Se não há categoria no formulário, verificar se já existe inscrição sem categoria
            tem_inscricao_sem_categoria = False
            if not categorias_selecionadas:
                for insc_existente in inscricoes_existentes:
                    if insc_existente.get("dados_form"):
                        try:
                            dados_existentes = json.loads(insc_existente["dados_form"]) if isinstance(insc_existente["dados_form"], str) else insc_existente["dados_form"]
                            if not dados_existentes.get("categoria"):
                                tem_inscricao_sem_categoria = True
                                break
                        except Exception:
                            # Se não conseguir parsear, considerar como sem categoria
                            tem_inscricao_sem_categoria = True
                            break
                    else:
                        # Se dados_form é None ou vazio, considerar como sem categoria
                        tem_inscricao_sem_categoria = True
                        break
            
            # Validar duplicatas
            if not categorias_selecionadas and tem_inscricao_sem_categoria:
                flash("Você já está inscrito neste evento.", "warning")
                return redirect(url_for("eventos_competicoes.disponiveis", academia_id=academia_id, aluno_id=aluno_id))
            
            # Filtrar categorias duplicadas, mas permitir criar inscrições para as válidas
            categorias_duplicadas = []
            categorias_validas = []
            if categorias_selecionadas:
                categorias_duplicadas = [cat for cat in categorias_selecionadas if cat in categorias_ja_inscritas]
                categorias_validas = [cat for cat in categorias_selecionadas if cat not in categorias_ja_inscritas]
                
                # Se todas as categorias selecionadas já estão inscritas, avisar e retornar
                if categorias_validas == [] and categorias_duplicadas:
                    flash(f"Você já está inscrito em todas as categorias selecionadas: {', '.join(categorias_duplicadas)}.", "warning")
                    return redirect(url_for("eventos_competicoes.disponiveis", academia_id=academia_id, aluno_id=aluno_id))
            
            inscricoes_criadas = 0
            categorias_criadas = []
            
            # Se não houver categoria selecionada, criar uma inscrição normal
            if not categorias_selecionadas:
                dados = dados_base.copy()
                cur.execute("""
                    INSERT INTO eventos_competicoes_inscricoes (evento_id, academia_id, aluno_id, usuario_inscricao_id, dados_form, inclusao_avulsa, status)
                    VALUES (%s, %s, %s, %s, %s, 0, 'confirmada')
                """, (evento_id, academia_id, aluno_id, current_user.id, json.dumps(dados, ensure_ascii=False)))
                inscricao_id = cur.lastrowid
                # Criar pagamento se evento tem taxa
                _criar_pagamento_inscricao(cur, inscricao_id, evento_id, academia_id)
                inscricoes_criadas = 1
            else:
                # Criar uma inscrição para cada categoria válida selecionada
                for categoria_nome in categorias_validas:
                    dados = dados_base.copy()
                    dados["categoria"] = categoria_nome
                    cur.execute("""
                        INSERT INTO eventos_competicoes_inscricoes (evento_id, academia_id, aluno_id, usuario_inscricao_id, dados_form, inclusao_avulsa, status)
                        VALUES (%s, %s, %s, %s, %s, 0, 'confirmada')
                    """, (evento_id, academia_id, aluno_id, current_user.id, json.dumps(dados, ensure_ascii=False)))
                    inscricao_id = cur.lastrowid
                    # Criar pagamento se evento tem taxa
                    _criar_pagamento_inscricao(cur, inscricao_id, evento_id, academia_id)
                    inscricoes_criadas += 1
                    categorias_criadas.append(categoria_nome)
                    categorias_ja_inscritas.add(categoria_nome)  # Adicionar à lista para evitar duplicatas na mesma requisição
            
            if inscricoes_criadas == 0:
                flash("Nenhuma inscrição foi criada. Verifique se há categorias válidas para inscrever.", "warning")
                return redirect(url_for("eventos_competicoes.disponiveis", academia_id=academia_id, aluno_id=aluno_id))
            
            # Atualizar dados do aluno conforme form (mapear form -> coluna DB)
            dados = dados_base.copy()  # Usar dados_base para atualização do aluno
            MAPEAMENTO_FORM_ALUNO = {
                "nome": "nome", "sexo": "sexo", "nome_pai": "nome_pai", "nome_mae": "nome_mae",
                "peso": "peso", "cpf": "cpf", "rg": "rg", "nacionalidade": "nacionalidade",
                "orgao_emissor": "orgao_emissor", "rg_data_emissao": "rg_data_emissao",
                "cep": "cep", "endereco": "rua", "numero": "numero", "bairro": "bairro",
                "cidade": "cidade", "estado": "estado", "complemento": "complemento",
                "email": "email", "observacoes": "observacoes", "zempo": "zempo",
                "responsavel_nome": "responsavel_nome", "responsavel_parentesco": "responsavel_parentesco",
                "responsavel_grau_parentesco": "responsavel_grau_parentesco",
                "responsavel_financeiro_nome": "responsavel_financeiro_nome",
                "responsavel_financeiro_cpf": "responsavel_financeiro_cpf",
                "telefone_celular": "tel_celular", "telefone_residencial": "tel_residencial",
                "telefone_comercial": "tel_comercial", "telefone_outro": "tel_outro",
                "ultimo_exame_faixa": "ultimo_exame_faixa", "data_cadastro_zempo": "data_cadastro_zempo",
                "graduacao_id": "graduacao_id", "TurmaID": "TurmaID", "data_nascimento": "data_nascimento",
            }
            for chave, val in dados.items():
                col = MAPEAMENTO_FORM_ALUNO.get(chave)
                if not col or chave in ("id_academia", "aluno_modalidade_ids", "foto"):
                    continue
                if col in ("graduacao_id", "TurmaID"):
                    try:
                        val = int(val) if val else None
                    except (ValueError, TypeError):
                        val = None
                elif col == "peso":
                    try:
                        val = float(val) if val else None
                    except (ValueError, TypeError):
                        val = None
                elif col in ("data_nascimento", "ultimo_exame_faixa", "rg_data_emissao", "data_cadastro_zempo"):
                    try:
                        from datetime import datetime as dt
                        val_str = str(val)[:10].strip()
                        if not val_str:
                            val = None
                        else:
                            # Tentar formato BR primeiro (dd/mm/yyyy), depois ISO (yyyy-mm-dd)
                            try:
                                val = dt.strptime(val_str, "%d/%m/%Y").date()
                            except ValueError:
                                try:
                                    val = dt.strptime(val_str, "%Y-%m-%d").date()
                                except ValueError:
                                    val = None
                    except Exception:
                        val = None
                if val is not None:
                    try:
                        cur.execute(f"UPDATE alunos SET `{col}` = %s WHERE id = %s", (val, aluno_id))
                    except Exception:
                        pass
            conn.commit()
            
            # Mensagem de sucesso informando quantas inscrições foram criadas
            if categorias_selecionadas:
                if len(categorias_selecionadas) > 1:
                    if categorias_duplicadas:
                        flash(f"{inscricoes_criadas} inscrição(ões) criada(s) com sucesso para as categorias: {', '.join(categorias_criadas)}. As categorias {', '.join(categorias_duplicadas)} já estavam inscritas.", "success")
                    else:
                        flash(f"{inscricoes_criadas} inscrições realizadas com sucesso! Uma para cada categoria selecionada.", "success")
                else:
                    flash("Inscrição realizada com sucesso!", "success")
            else:
                flash("Inscrição realizada com sucesso!", "success")
            return redirect(url_for("eventos_competicoes.disponiveis", academia_id=academia_id, aluno_id=aluno_id))

        # GET: montar valores iniciais do aluno (mapear col DB -> form)
        REV_MAP = {"rua": "endereco", "tel_celular": "telefone_celular", "tel_residencial": "telefone_residencial",
                   "tel_comercial": "telefone_comercial", "tel_outro": "telefone_outro", "telefone": "telefone_celular"}
        valores_iniciais = {}
        for c in campos_form:
            chave = c["campo_chave"]
            v = aluno.get(chave)
            if v is None and chave in ("endereco",):
                v = aluno.get("rua")
            elif v is None and chave == "telefone_celular":
                v = aluno.get("tel_celular") or aluno.get("telefone")
            elif v is None and chave in ("telefone_residencial", "telefone_comercial", "telefone_outro"):
                v = aluno.get("tel_" + chave.split("_")[1])
            if hasattr(v, "strftime"):
                # Para input type="text" com formato BR, converter para BR
                v = v.strftime("%d/%m/%Y") if v else ""
            elif chave == "id_academia":
                # Se vier ID, usar o ID da academia atual
                v = academia_id or v
            elif chave == "graduacao_id" and v:
                # Manter ID para o value (o select já mostra o nome)
                pass
            elif chave == "TurmaID" and v:
                # Manter ID para o value (o select já mostra o nome)
                pass
            valores_iniciais[chave] = v or ""

        # Buscar nome da academia para exibição
        academia_nome = None
        if academia_id:
            cur.execute("SELECT nome FROM academias WHERE id = %s", (academia_id,))
            ac_row = cur.fetchone()
            academia_nome = ac_row["nome"] if ac_row else None
        
        return render_template("eventos_competicoes/inscrever.html",
            evento=ev, aluno=aluno, academia_id=academia_id, academia_nome=academia_nome,
            campos_form=campos_form, valores_iniciais=valores_iniciais,
            graduacoes=graduacoes, turmas=turmas, professores=professores,
            categorias_disponiveis=categorias_disponiveis,
            back_url=request.referrer or url_for("painel.home"))
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/<int:evento_id>/pagamentos")
@login_required
def pagamentos_academias(evento_id):
    """Associação: gerencia pagamentos recebidos das academias."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se colunas de taxa existem
        tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
        if tem_coluna_taxa:
            cur.execute("""
                SELECT id, nome, data_fim, tem_taxa, valor_taxa_sugerido
                FROM eventos_competicoes WHERE id = %s AND id_associacao = %s
            """, (evento_id, id_assoc))
        else:
            cur.execute("""
                SELECT id, nome, data_fim, 0 as tem_taxa, NULL as valor_taxa_sugerido
                FROM eventos_competicoes WHERE id = %s AND id_associacao = %s
            """, (evento_id, id_assoc))
        evento = cur.fetchone()
        if not evento:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))
        
        # Verificar se tabelas de pagamentos existem
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
        
        # Buscar academias que enviaram inscrições
        cur.execute("""
            SELECT DISTINCT i.academia_id, ac.nome as academia_nome,
                   COUNT(*) as total_inscricoes
            FROM eventos_competicoes_inscricoes i
            INNER JOIN academias ac ON ac.id = i.academia_id
            WHERE i.evento_id = %s AND i.status = 'enviada'
            GROUP BY i.academia_id, ac.nome
            ORDER BY ac.nome
        """, (evento_id,))
        academias_com_inscricoes = cur.fetchall()
        
        pagamentos = []
        # Valor da associação para calcular o que a academia deve pagar
        valor_taxa_associacao = float(evento.get("valor_taxa_sugerido", 0) or 0)
        
        if tabela_pagamentos_existe:
            # Buscar pagamentos das academias
            cur.execute("""
                SELECT ap.id, ap.academia_id, 
                       COALESCE(ap.valor_total_esperado, 0) as valor_total_esperado,
                       COALESCE(ap.valor_pago, 0) as valor_pago,
                       COALESCE(ap.valor_pendente, 0) as valor_pendente,
                       ap.status
                FROM eventos_competicoes_academia_pagamentos ap
                WHERE ap.evento_id = %s
            """, (evento_id,))
            pagamentos_dict = {p["academia_id"]: p for p in cur.fetchall()}
            
            # Buscar histórico de abatimentos para cada pagamento
            try:
                cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos_abatimentos'")
                tabela_abatimentos_existe = cur.fetchone() is not None
            except Exception:
                tabela_abatimentos_existe = False
            
            # Combinar dados de academias com inscrições e pagamentos
            for ac in academias_com_inscricoes:
                try:
                    pagamento_data = pagamentos_dict.get(ac.get("academia_id"))
                    if pagamento_data:
                        pag = pagamento_data.copy()
                        pag["academia_nome"] = ac.get("academia_nome", "")
                        total_insc = ac.get("total_inscricoes", 0) or 0
                        pag["total_inscricoes"] = total_insc
                        
                        # Garantir que valores sejam float
                        pag["valor_total_esperado"] = float(pag.get("valor_total_esperado") or 0)
                        pag["valor_pago"] = float(pag.get("valor_pago") or 0)
                        pag["valor_pendente"] = float(pag.get("valor_pendente") or 0)
                        
                        # Recalcular valor_total_esperado baseado no valor da associação
                        valor_total_esperado_correto = float(valor_taxa_associacao * total_insc)
                        
                        # Se o valor está incorreto ou zerado, atualizar o registro
                        valor_atual = float(pag.get("valor_total_esperado", 0) or 0)
                        valor_pago_atual = float(pag.get("valor_pago", 0) or 0)
                        
                        # Atualizar se valor está zerado ou diferente (com tolerância)
                        # Só recalcular se valor_taxa_associacao > 0 (evento tem taxa)
                        if valor_taxa_associacao > 0 and (valor_atual == 0 or abs(valor_atual - valor_total_esperado_correto) > 0.01):
                            # Recalcular valor_pendente baseado no valor correto
                            novo_valor_pendente = valor_total_esperado_correto - valor_pago_atual
                            
                            # Atualizar status se necessário
                            if novo_valor_pendente <= 0:
                                novo_status = "quitado"
                            elif valor_pago_atual > 0:
                                novo_status = "parcial"
                            else:
                                novo_status = "pendente"
                            
                            # Atualizar o registro no banco
                            cur.execute("""
                                UPDATE eventos_competicoes_academia_pagamentos
                                SET valor_total_esperado = %s, valor_pendente = %s, status = %s
                                WHERE id = %s
                            """, (valor_total_esperado_correto, novo_valor_pendente, novo_status, pag["id"]))
                            conn.commit()
                            
                            # Atualizar o valor no objeto pag
                            pag["valor_total_esperado"] = valor_total_esperado_correto
                            pag["valor_pago"] = valor_pago_atual
                            pag["valor_pendente"] = novo_valor_pendente
                            pag["status"] = novo_status
                        else:
                            # Mesmo que o valor esteja correto, garantir que valor_pendente está correto
                            valor_pendente_atual = float(pag.get("valor_pendente", 0) or 0)
                            valor_pendente_correto = valor_total_esperado_correto - valor_pago_atual
                            # Sempre recalcular e atualizar valor_pendente para garantir consistência
                            pag["valor_pendente"] = valor_pendente_correto
                            # Atualizar no banco também
                            cur.execute("""
                                UPDATE eventos_competicoes_academia_pagamentos
                                SET valor_pendente = %s
                                WHERE id = %s
                            """, (valor_pendente_correto, pag["id"]))
                            conn.commit()
                        
                        # Buscar abatimentos se existir tabela
                        if tabela_abatimentos_existe and pag.get("id"):
                            try:
                                cur.execute("""
                                    SELECT id, valor, data_abatimento, observacoes
                                    FROM eventos_competicoes_academia_pagamentos_abatimentos
                                    WHERE pagamento_id = %s
                                    ORDER BY data_abatimento DESC
                                """, (pag["id"],))
                                pag["abatimentos"] = cur.fetchall()
                            except Exception:
                                pag["abatimentos"] = []
                        else:
                            pag["abatimentos"] = []
                        
                        pagamentos.append(pag)
                    else:
                        # Academia enviou inscrições mas não tem registro de pagamento ainda
                        # Usar valor da associação (valor_taxa_sugerido) para calcular o que a academia deve pagar
                        # valor_taxa_associacao já foi definido antes do loop
                        valor_total_esperado = valor_taxa_associacao * (ac.get("total_inscricoes", 0) or 0)
                        
                        # Criar registro de pagamento automaticamente se evento tem taxa
                        tem_taxa_val = evento.get("tem_taxa", 0)
                        tem_taxa_int = int(tem_taxa_val) if tem_taxa_val else 0
                        # Criar sempre que houver valor esperado, mesmo se evento não tem taxa configurada
                        if valor_total_esperado > 0:
                            try:
                                cur.execute("""
                                    INSERT INTO eventos_competicoes_academia_pagamentos
                                    (evento_id, academia_id, valor_total_esperado, valor_pendente, status)
                                    VALUES (%s, %s, %s, %s, 'pendente')
                                """, (evento_id, ac.get("academia_id"), valor_total_esperado, valor_total_esperado))
                                pagamento_id = cur.lastrowid
                                conn.commit()
                                
                                # Buscar o registro criado
                                cur.execute("""
                                    SELECT id, valor_total_esperado, valor_pago, valor_pendente, status
                                    FROM eventos_competicoes_academia_pagamentos
                                    WHERE id = %s
                                """, (pagamento_id,))
                                pag_data = cur.fetchone()
                                
                                if pag_data:
                                    pag = {
                                        "id": pag_data["id"],
                                        "academia_id": ac.get("academia_id"),
                                        "academia_nome": ac.get("academia_nome", ""),
                                        "total_inscricoes": ac.get("total_inscricoes", 0) or 0,
                                        "valor_total_esperado": float(pag_data.get("valor_total_esperado") or 0),
                                        "valor_pago": float(pag_data.get("valor_pago") or 0),
                                        "valor_pendente": float(pag_data.get("valor_pendente") or 0),
                                        "status": pag_data.get("status", "pendente"),
                                        "abatimentos": []
                                    }
                                else:
                                    # Se não encontrou o registro, criar sem ID
                                    pag = {
                                        "id": None,
                                        "academia_id": ac.get("academia_id"),
                                        "academia_nome": ac.get("academia_nome", ""),
                                        "total_inscricoes": ac.get("total_inscricoes", 0) or 0,
                                        "valor_total_esperado": float(valor_total_esperado),
                                        "valor_pago": 0.0,
                                        "valor_pendente": float(valor_total_esperado),
                                        "status": "pendente",
                                        "abatimentos": []
                                    }
                            except Exception as e:
                                # Se houver erro ao criar, usar registro sem ID
                                try:
                                    import traceback
                                    error_msg = f"Erro ao criar registro de pagamento para academia {ac.get('academia_id')}: {e}\n{traceback.format_exc()}"
                                    current_app.logger.error(error_msg)
                                    print(error_msg)
                                except:
                                    pass
                                pag = {
                                    "id": None,
                                    "academia_id": ac.get("academia_id"),
                                    "academia_nome": ac.get("academia_nome", ""),
                                    "total_inscricoes": ac.get("total_inscricoes", 0) or 0,
                                    "valor_total_esperado": float(valor_total_esperado),
                                    "valor_pago": 0.0,
                                    "valor_pendente": float(valor_total_esperado),
                                    "status": "pendente",
                                    "abatimentos": []
                                }
                        else:
                            # Se não tem valor esperado, ainda assim criar registro sem ID para exibição
                            pag = {
                                "id": None,
                                "academia_id": ac.get("academia_id"),
                                "academia_nome": ac.get("academia_nome", ""),
                                "total_inscricoes": ac.get("total_inscricoes", 0),
                                "valor_total_esperado": valor_total_esperado,
                                "valor_pago": 0,
                                "valor_pendente": valor_total_esperado,
                                "status": "pendente",
                                "abatimentos": []
                            }
                        pagamentos.append(pag)
                except Exception as e:
                    # Se houver erro ao processar esta academia, pular
                    try:
                        current_app.logger.error(f"Erro ao processar academia {ac.get('academia_id')}: {e}")
                    except:
                        pass
                    continue
        else:
            # Se tabela não existe, mostrar apenas academias com inscrições
            # Usar valor da associação (valor_taxa_sugerido) para calcular o que a academia deve pagar
            valor_taxa_associacao = float(evento.get("valor_taxa_sugerido", 0) or 0)
            for ac in academias_com_inscricoes:
                try:
                    total_insc = ac.get("total_inscricoes", 0) or 0
                    valor_total_esperado = float(valor_taxa_associacao * total_insc)
                    
                    pag = {
                        "id": None,
                        "academia_id": ac.get("academia_id"),
                        "academia_nome": ac.get("academia_nome", ""),
                        "total_inscricoes": total_insc,
                        "valor_total_esperado": valor_total_esperado,
                        "valor_pago": 0.0,
                        "valor_pendente": valor_total_esperado,
                        "status": "pendente",
                        "abatimentos": []
                    }
                    pagamentos.append(pag)
                except Exception as e:
                    # Se houver erro, pular esta academia
                    try:
                        current_app.logger.error(f"Erro ao processar academia {ac.get('academia_id')}: {e}")
                    except:
                        pass
                    continue
        
        # Garantir que todos os valores estão definidos antes de passar para o template
        if not isinstance(pagamentos, list):
            pagamentos = []
        
        for pag in pagamentos:
            if not isinstance(pag, dict):
                continue
            pag.setdefault("id", None)
            pag.setdefault("academia_id", None)
            pag.setdefault("academia_nome", "")
            pag.setdefault("total_inscricoes", 0)
            
            # Garantir que valores financeiros sejam sempre float
            valor_total_esperado_raw = pag.get("valor_total_esperado")
            if valor_total_esperado_raw is None:
                # Se não tem valor, calcular baseado no valor da associação
                total_insc = pag.get("total_inscricoes", 0) or 0
                pag["valor_total_esperado"] = float(valor_taxa_associacao * total_insc)
            else:
                pag["valor_total_esperado"] = float(valor_total_esperado_raw or 0)
            
            valor_pago_raw = pag.get("valor_pago")
            pag["valor_pago"] = float(valor_pago_raw or 0) if valor_pago_raw is not None else 0.0
            
            valor_pendente_raw = pag.get("valor_pendente")
            if valor_pendente_raw is None:
                # Recalcular se não existe
                pag["valor_pendente"] = pag["valor_total_esperado"] - pag["valor_pago"]
            else:
                pag["valor_pendente"] = float(valor_pendente_raw or 0)
            
            pag.setdefault("status", "pendente")
            pag.setdefault("abatimentos", [])
        
        # Garantir que evento tem todas as propriedades necessárias
        if not isinstance(evento, dict):
            flash("Erro ao carregar dados do evento.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))
        
        evento.setdefault("tem_taxa", 0)
        if evento.get("tem_taxa"):
            evento["tem_taxa"] = int(evento["tem_taxa"])
        else:
            evento["tem_taxa"] = 0
        
        # Calcular totais financeiros
        total_esperado = sum(float(p.get("valor_total_esperado", 0) or 0) for p in pagamentos)
        total_pago = sum(float(p.get("valor_pago", 0) or 0) for p in pagamentos)
        total_pendente = sum(float(p.get("valor_pendente", 0) or 0) for p in pagamentos)
        
        return render_template("eventos_competicoes/pagamentos_academias.html",
            evento=evento, pagamentos=pagamentos,
            total_esperado=total_esperado, total_pago=total_pago, total_pendente=total_pendente,
            back_url=url_for("eventos_competicoes.lista"))
    except Exception as e:
        # Log do erro para debug
        try:
            current_app.logger.error(f"Erro em pagamentos_academias: {e}")
            import traceback
            current_app.logger.error(traceback.format_exc())
        except:
            pass
        flash("Erro ao carregar página de pagamentos. Verifique os logs.", "danger")
        return redirect(url_for("eventos_competicoes.lista"))
    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass


@bp_eventos_competicoes.route("/<int:evento_id>/pagamentos/criar-registro", methods=["POST"])
@login_required
def criar_registro_pagamento(evento_id):
    """Associação cria registro de pagamento manualmente para uma academia."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    academia_id = request.form.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se evento pertence à associação e buscar valor_taxa_sugerido
        tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
        if tem_coluna_taxa:
            cur.execute("SELECT id, valor_taxa_sugerido FROM eventos_competicoes WHERE id = %s AND id_associacao = %s", (evento_id, id_assoc))
        else:
            cur.execute("SELECT id, NULL as valor_taxa_sugerido FROM eventos_competicoes WHERE id = %s AND id_associacao = %s", (evento_id, id_assoc))
        evento = cur.fetchone()
        if not evento:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))
        
        # Verificar se tabela existe
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
            tabela_existe = cur.fetchone() is not None
        except Exception:
            tabela_existe = False
        
        if not tabela_existe:
            flash("Sistema de pagamentos não está disponível. Execute a migration primeiro.", "warning")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Verificar se já existe registro
        cur.execute("""
            SELECT id FROM eventos_competicoes_academia_pagamentos
            WHERE evento_id = %s AND academia_id = %s
        """, (evento_id, academia_id))
        if cur.fetchone():
            flash("Registro de pagamento já existe para esta academia.", "info")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Usar valor da associação (valor_taxa_sugerido) para calcular o que a academia deve pagar
        valor_taxa_associacao = float(evento.get("valor_taxa_sugerido", 0) or 0)
        
        cur.execute("""
            SELECT COUNT(*) as total FROM eventos_competicoes_inscricoes
            WHERE evento_id = %s AND academia_id = %s AND status = 'enviada'
        """, (evento_id, academia_id))
        total_result = cur.fetchone()
        total_inscricoes = total_result.get("total", 0) or 0 if total_result else 0
        
        valor_total_esperado = valor_taxa_associacao * total_inscricoes
        
        if valor_total_esperado <= 0:
            flash("Não é possível criar registro: valor total esperado é zero.", "warning")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Criar registro
        cur.execute("""
            INSERT INTO eventos_competicoes_academia_pagamentos
            (evento_id, academia_id, valor_total_esperado, valor_pendente, status)
            VALUES (%s, %s, %s, %s, 'pendente')
        """, (evento_id, academia_id, valor_total_esperado, valor_total_esperado))
        conn.commit()
        flash("Registro de pagamento criado com sucesso.", "success")
    except Exception as e:
        try:
            import traceback
            current_app.logger.error(f"Erro ao criar registro de pagamento: {e}\n{traceback.format_exc()}")
        except:
            pass
        flash(f"Erro ao criar registro: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))


@bp_eventos_competicoes.route("/<int:evento_id>/pagamentos/<int:pagamento_id>/abatimento", methods=["POST"])
@login_required
def registrar_abatimento(evento_id, pagamento_id):
    """Associação registra abatimento de pagamento de uma academia."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    valor_abatimento = request.form.get("valor_abatimento", type=float)
    observacoes = request.form.get("observacoes", "").strip()
    
    if not valor_abatimento or valor_abatimento <= 0:
        flash("Informe um valor válido para o abatimento.", "danger")
        return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se pagamento pertence ao evento da associação
        cur.execute("""
            SELECT ap.id, ap.valor_pendente, ap.valor_pago, ec.id_associacao
            FROM eventos_competicoes_academia_pagamentos ap
            INNER JOIN eventos_competicoes ec ON ec.id = ap.evento_id
            WHERE ap.id = %s AND ap.evento_id = %s AND ec.id_associacao = %s
        """, (pagamento_id, evento_id, id_assoc))
        pagamento = cur.fetchone()
        
        if not pagamento:
            flash("Pagamento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        valor_pendente_atual = float(pagamento.get("valor_pendente", 0) or 0)
        valor_pago_atual = float(pagamento.get("valor_pago", 0) or 0)
        
        if valor_abatimento > valor_pendente_atual:
            flash(f"O valor do abatimento não pode ser maior que o pendente (R$ {valor_pendente_atual:.2f}).", "danger")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Registrar abatimento
        cur.execute("""
            INSERT INTO eventos_competicoes_academia_pagamentos_abatimentos
            (pagamento_id, valor, observacoes, criado_por_usuario_id)
            VALUES (%s, %s, %s, %s)
        """, (pagamento_id, valor_abatimento, observacoes or None, current_user.id))
        
        # Atualizar valores do pagamento
        novo_valor_pago = valor_pago_atual + valor_abatimento
        novo_valor_pendente = valor_pendente_atual - valor_abatimento
        novo_status = "quitado" if novo_valor_pendente <= 0 else ("parcial" if novo_valor_pago > 0 else "pendente")
        
        cur.execute("""
            UPDATE eventos_competicoes_academia_pagamentos
            SET valor_pago = %s, valor_pendente = %s, status = %s
            WHERE id = %s
        """, (novo_valor_pago, novo_valor_pendente, novo_status, pagamento_id))
        
        conn.commit()
        flash(f"Abatimento de R$ {valor_abatimento:.2f} registrado com sucesso.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))


@bp_eventos_competicoes.route("/<int:evento_id>/pagamentos/<int:pagamento_id>/confirmar", methods=["POST"])
@login_required
def confirmar_pagamento_academia(evento_id, pagamento_id):
    """Associação confirma pagamento geral de uma academia (marca todas as inscrições como pagas pela associação)."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se tabelas de pagamentos existem
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
            tabela_pagamentos_inscricoes_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
            tabela_pagamentos_inscricoes_existe = False
        
        if not tabela_pagamentos_existe or not tabela_pagamentos_inscricoes_existe:
            flash("Sistema de pagamentos não está disponível.", "warning")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Verificar se pagamento pertence ao evento da associação
        cur.execute("""
            SELECT ap.id, ap.academia_id, ec.id_associacao
            FROM eventos_competicoes_academia_pagamentos ap
            INNER JOIN eventos_competicoes ec ON ec.id = ap.evento_id
            WHERE ap.id = %s AND ap.evento_id = %s AND ec.id_associacao = %s
        """, (pagamento_id, evento_id, id_assoc))
        pagamento = cur.fetchone()
        
        if not pagamento:
            flash("Pagamento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Marcar todas as inscrições enviadas desta academia como pagas pela associação
        cur.execute("""
            UPDATE eventos_competicoes_inscricoes_pagamentos p
            INNER JOIN eventos_competicoes_inscricoes i ON i.id = p.inscricao_id
            SET p.pago_associacao = 1, p.data_pagamento_associacao = NOW()
            WHERE i.evento_id = %s AND i.academia_id = %s AND i.status = 'enviada'
        """, (evento_id, pagamento["academia_id"]))
        
        conn.commit()
        flash("Pagamento confirmado. Todas as inscrições desta academia foram marcadas como pagas.", "success")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))


@bp_eventos_competicoes.route("/<int:evento_id>/pagamentos/<int:pagamento_id>/cancelar", methods=["POST"])
@login_required
def cancelar_pagamento_academia(evento_id, pagamento_id):
    """Associação cancela/reverte pagamento de uma academia (reverte último abatimento ou pagamento total)."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se tabelas de pagamentos existem
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos_abatimentos'")
            tabela_abatimentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
            tabela_abatimentos_existe = False
        
        if not tabela_pagamentos_existe:
            flash("Sistema de pagamentos não está disponível.", "warning")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Verificar se pagamento pertence ao evento da associação
        cur.execute("""
            SELECT ap.id, ap.academia_id, ap.valor_total_esperado, ap.valor_pago, ap.valor_pendente, ap.status, ec.id_associacao
            FROM eventos_competicoes_academia_pagamentos ap
            INNER JOIN eventos_competicoes ec ON ec.id = ap.evento_id
            WHERE ap.id = %s AND ap.evento_id = %s AND ec.id_associacao = %s
        """, (pagamento_id, evento_id, id_assoc))
        pagamento = cur.fetchone()
        
        if not pagamento:
            flash("Pagamento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        valor_pago_atual = float(pagamento.get("valor_pago", 0) or 0)
        valor_total_esperado = float(pagamento.get("valor_total_esperado", 0) or 0)
        
        if valor_pago_atual <= 0:
            flash("Não há pagamento para cancelar.", "warning")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Se existe tabela de abatimentos, tentar remover o último abatimento
        if tabela_abatimentos_existe:
            cur.execute("""
                SELECT id, valor FROM eventos_competicoes_academia_pagamentos_abatimentos
                WHERE pagamento_id = %s
                ORDER BY data_abatimento DESC, id DESC
                LIMIT 1
            """, (pagamento_id,))
            ultimo_abatimento = cur.fetchone()
            
            if ultimo_abatimento:
                # Remover o último abatimento
                valor_abatimento = float(ultimo_abatimento.get("valor", 0) or 0)
                novo_valor_pago = valor_pago_atual - valor_abatimento
                novo_valor_pendente = valor_total_esperado - novo_valor_pago
                
                cur.execute("DELETE FROM eventos_competicoes_academia_pagamentos_abatimentos WHERE id = %s", (ultimo_abatimento["id"],))
                
                # Atualizar status do pagamento
                if novo_valor_pendente <= 0:
                    novo_status = "quitado"
                elif novo_valor_pago > 0:
                    novo_status = "parcial"
                else:
                    novo_status = "pendente"
                
                cur.execute("""
                    UPDATE eventos_competicoes_academia_pagamentos
                    SET valor_pago = %s, valor_pendente = %s, status = %s
                    WHERE id = %s
                """, (novo_valor_pago, novo_valor_pendente, novo_status, pagamento_id))
                
                conn.commit()
                flash(f"Último abatimento de R$ {valor_abatimento:.2f} foi cancelado. Pagamento revertido.", "success")
            else:
                # Se não há abatimentos, reverter pagamento total para zero
                cur.execute("""
                    UPDATE eventos_competicoes_academia_pagamentos
                    SET valor_pago = 0, valor_pendente = %s, status = 'pendente'
                    WHERE id = %s
                """, (valor_total_esperado, pagamento_id))
                conn.commit()
                flash("Pagamento cancelado. Status revertido para pendente.", "success")
        else:
            # Se não há tabela de abatimentos, reverter pagamento total para zero
            cur.execute("""
                UPDATE eventos_competicoes_academia_pagamentos
                SET valor_pago = 0, valor_pendente = %s, status = 'pendente'
                WHERE id = %s
            """, (valor_total_esperado, pagamento_id))
            conn.commit()
            flash("Pagamento cancelado. Status revertido para pendente.", "success")
    except Exception as e:
        try:
            import traceback
            current_app.logger.error(f"Erro ao cancelar pagamento: {e}\n{traceback.format_exc()}")
        except:
            pass
        flash(f"Erro ao cancelar pagamento: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))


@bp_eventos_competicoes.route("/<int:evento_id>/pagamentos/<int:pagamento_id>/abatimentos/<int:abatimento_id>/cancelar", methods=["POST"])
@login_required
def cancelar_abatimento(evento_id, pagamento_id, abatimento_id):
    """Associação cancela um abatimento específico."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se tabelas de pagamentos existem
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos_abatimentos'")
            tabela_abatimentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
            tabela_abatimentos_existe = False
        
        if not tabela_pagamentos_existe or not tabela_abatimentos_existe:
            flash("Sistema de pagamentos não está disponível.", "warning")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Verificar se abatimento pertence ao pagamento e evento da associação
        cur.execute("""
            SELECT ab.id, ab.pagamento_id, ab.valor, ap.valor_pago, ap.valor_pendente, ap.valor_total_esperado, ec.id_associacao
            FROM eventos_competicoes_academia_pagamentos_abatimentos ab
            INNER JOIN eventos_competicoes_academia_pagamentos ap ON ap.id = ab.pagamento_id
            INNER JOIN eventos_competicoes ec ON ec.id = ap.evento_id
            WHERE ab.id = %s AND ab.pagamento_id = %s AND ap.evento_id = %s AND ec.id_associacao = %s
        """, (abatimento_id, pagamento_id, evento_id, id_assoc))
        abatimento = cur.fetchone()
        
        if not abatimento:
            flash("Abatimento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        valor_abatimento = float(abatimento.get("valor", 0) or 0)
        valor_pago_atual = float(abatimento.get("valor_pago", 0) or 0)
        valor_total_esperado = float(abatimento.get("valor_total_esperado", 0) or 0)
        
        # Remover o abatimento
        cur.execute("DELETE FROM eventos_competicoes_academia_pagamentos_abatimentos WHERE id = %s", (abatimento_id,))
        
        # Recalcular valores do pagamento
        novo_valor_pago = valor_pago_atual - valor_abatimento
        novo_valor_pendente = valor_total_esperado - novo_valor_pago
        
        # Atualizar status do pagamento
        if novo_valor_pendente <= 0:
            novo_status = "quitado"
        elif novo_valor_pago > 0:
            novo_status = "parcial"
        else:
            novo_status = "pendente"
        
        cur.execute("""
            UPDATE eventos_competicoes_academia_pagamentos
            SET valor_pago = %s, valor_pendente = %s, status = %s
            WHERE id = %s
        """, (novo_valor_pago, novo_valor_pendente, novo_status, pagamento_id))
        
        conn.commit()
        flash(f"Abatimento de R$ {valor_abatimento:.2f} foi cancelado. Valores recalculados.", "success")
    except Exception as e:
        try:
            import traceback
            current_app.logger.error(f"Erro ao cancelar abatimento: {e}\n{traceback.format_exc()}")
        except:
            pass
        flash(f"Erro ao cancelar abatimento: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))


@bp_eventos_competicoes.route("/<int:evento_id>/cancelar-envio", methods=["POST"])
@login_required
def cancelar_envio(evento_id):
    """Associação cancela envio de inscrições de uma academia, permitindo que ela edite novamente."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    academia_id = request.form.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se evento pertence à associação
        cur.execute("SELECT id FROM eventos_competicoes WHERE id = %s AND id_associacao = %s", (evento_id, id_assoc))
        evento = cur.fetchone()
        if not evento:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))
        
        # Verificar se há inscrições enviadas desta academia
        cur.execute("""
            SELECT COUNT(*) as total FROM eventos_competicoes_inscricoes
            WHERE evento_id = %s AND academia_id = %s AND status = 'enviada'
        """, (evento_id, academia_id))
        result = cur.fetchone()
        total_enviadas = result.get("total", 0) or 0 if result else 0
        
        if total_enviadas == 0:
            flash("Não há inscrições enviadas para cancelar.", "info")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Reverter status de 'enviada' para 'confirmada'
        cur.execute("""
            UPDATE eventos_competicoes_inscricoes
            SET status = 'confirmada', data_envio = NULL
            WHERE evento_id = %s AND academia_id = %s AND status = 'enviada'
        """, (evento_id, academia_id))
        
        # Se evento tem taxa, atualizar registro de pagamento (recalcular valores)
        tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
        if tem_coluna_taxa:
            cur.execute("SELECT valor_taxa_sugerido FROM eventos_competicoes WHERE id = %s", (evento_id,))
            evento_data = cur.fetchone()
            valor_taxa_associacao = float(evento_data.get("valor_taxa_sugerido", 0) or 0) if evento_data else 0
            
            if valor_taxa_associacao > 0:
                try:
                    cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
                    tabela_pagamentos_existe = cur.fetchone() is not None
                except Exception:
                    tabela_pagamentos_existe = False
                
                if tabela_pagamentos_existe:
                    # Contar inscrições confirmadas (não enviadas)
                    cur.execute("""
                        SELECT COUNT(*) as total_confirmadas
                        FROM eventos_competicoes_inscricoes
                        WHERE evento_id = %s AND academia_id = %s AND status = 'confirmada'
                    """, (evento_id, academia_id))
                    total_confirmadas = cur.fetchone().get("total_confirmadas", 0) or 0
                    novo_valor_total_esperado = valor_taxa_associacao * total_confirmadas
                    
                    # Atualizar registro de pagamento
                    cur.execute("""
                        SELECT id, valor_pago FROM eventos_competicoes_academia_pagamentos
                        WHERE evento_id = %s AND academia_id = %s
                    """, (evento_id, academia_id))
                    pagamento = cur.fetchone()
                    
                    if pagamento:
                        valor_pago_atual = float(pagamento.get("valor_pago", 0) or 0)
                        novo_valor_pendente = novo_valor_total_esperado - valor_pago_atual
                        
                        if novo_valor_pendente <= 0:
                            novo_status = "quitado"
                        elif valor_pago_atual > 0:
                            novo_status = "parcial"
                        else:
                            novo_status = "pendente"
                        
                        cur.execute("""
                            UPDATE eventos_competicoes_academia_pagamentos
                            SET valor_total_esperado = %s, valor_pendente = %s, status = %s
                            WHERE id = %s
                        """, (novo_valor_total_esperado, novo_valor_pendente, novo_status, pagamento["id"]))
        conn.commit()
        flash(f"Envio cancelado. {total_enviadas} inscrição(ões) voltaram ao status 'confirmada'. A academia pode editar novamente.", "success")
    except Exception as e:
        try:
            import traceback
            current_app.logger.error(f"Erro ao cancelar envio: {e}\n{traceback.format_exc()}")
        except:
            pass
        flash(f"Erro ao cancelar envio: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))


@bp_eventos_competicoes.route("/<int:evento_id>/cancelar-inscricoes-academia", methods=["POST"])
@login_required
def cancelar_inscricoes_academia(evento_id):
    """Associação cancela todas as inscrições de uma academia específica."""
    if session.get("modo_painel") != "associacao":
        flash("Disponível apenas em modo associação.", "danger")
        return redirect(url_for("painel.home"))
    if not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        flash("Selecione a associação.", "warning")
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    academia_id = request.form.get("academia_id", type=int)
    if not academia_id:
        flash("Academia não informada.", "danger")
        return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se evento pertence à associação
        cur.execute("SELECT id FROM eventos_competicoes WHERE id = %s AND id_associacao = %s", (evento_id, id_assoc))
        evento = cur.fetchone()
        if not evento:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))
        
        # Verificar se academia pertence à associação
        cur.execute("SELECT id, nome FROM academias WHERE id = %s AND id_associacao = %s", (academia_id, id_assoc))
        academia = cur.fetchone()
        if not academia:
            flash("Academia não encontrada.", "danger")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Contar inscrições antes de deletar
        cur.execute("""
            SELECT COUNT(*) as total FROM eventos_competicoes_inscricoes
            WHERE evento_id = %s AND academia_id = %s
        """, (evento_id, academia_id))
        total_inscricoes = cur.fetchone().get("total", 0) or 0
        
        if total_inscricoes == 0:
            flash("Não há inscrições para cancelar.", "info")
            return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
        
        # Deletar todas as inscrições da academia (cascade vai deletar pagamentos relacionados)
        cur.execute("""
            DELETE FROM eventos_competicoes_inscricoes
            WHERE evento_id = %s AND academia_id = %s
        """, (evento_id, academia_id))
        
        # Deletar registro de pagamento da academia se existir
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
        
        if tabela_pagamentos_existe:
            cur.execute("""
                DELETE FROM eventos_competicoes_academia_pagamentos
                WHERE evento_id = %s AND academia_id = %s
            """, (evento_id, academia_id))
        
        conn.commit()
        flash(f"Todas as inscrições da academia {academia['nome']} foram canceladas. {total_inscricoes} inscrição(ões) removida(s).", "success")
    except Exception as e:
        try:
            import traceback
            current_app.logger.error(f"Erro ao cancelar inscrições da academia: {e}\n{traceback.format_exc()}")
        except:
            pass
        flash(f"Erro ao cancelar inscrições: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))


@bp_eventos_competicoes.route("/<int:evento_id>/consolidar")
@login_required
def consolidar(evento_id):
    """Associação: lista academias com alunos inscritos, botão exportar PDF/Excel."""
    if session.get("modo_painel") != "associacao" or not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        return redirect(url_for("associacao.gerenciamento_associacao"))
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se evento tem taxa e se há valores pendentes
        tem_coluna_taxa, _ = _colunas_taxa_existem(cur)
        if tem_coluna_taxa:
            cur.execute("SELECT id, nome, tem_taxa FROM eventos_competicoes WHERE id = %s AND id_associacao = %s", (evento_id, id_assoc))
        else:
            cur.execute("SELECT id, nome, 0 as tem_taxa FROM eventos_competicoes WHERE id = %s AND id_associacao = %s", (evento_id, id_assoc))
        evento_check = cur.fetchone()
        
        if evento_check and evento_check.get("tem_taxa"):
            # Verificar se há valores pendentes
            try:
                cur.execute("SHOW TABLES LIKE 'eventos_competicoes_academia_pagamentos'")
                tabela_pagamentos_existe = cur.fetchone() is not None
            except Exception:
                tabela_pagamentos_existe = False
            
            if tabela_pagamentos_existe:
                cur.execute("""
                    SELECT SUM(valor_pendente) as total_pendente
                    FROM eventos_competicoes_academia_pagamentos
                    WHERE evento_id = %s
                """, (evento_id,))
                result = cur.fetchone()
                total_pendente = float(result.get("total_pendente", 0) or 0) if result else 0
                
                if total_pendente > 0:
                    flash(f"Não é possível consolidar. Ainda há R$ {total_pendente:.2f} em valores pendentes. Quite todos os pagamentos antes de consolidar.", "warning")
                    return redirect(url_for("eventos_competicoes.pagamentos_academias", evento_id=evento_id))
    finally:
        cur.close()
        conn.close()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT ec.id, ec.nome, ec.data_fim, ec.id_formulario
            FROM eventos_competicoes ec WHERE ec.id = %s AND ec.id_associacao = %s
        """, (evento_id, id_assoc))
        ev = cur.fetchone()
        if not ev:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))

        # Buscar academias com inscrições
        cur.execute("""
            SELECT DISTINCT ac.id as academia_id, ac.nome as academia_nome
            FROM academias ac
            INNER JOIN eventos_competicoes_adesao ea ON ea.academia_id = ac.id AND ea.evento_id = %s AND ea.aderiu = 1
            INNER JOIN eventos_competicoes_inscricoes i ON i.academia_id = ac.id AND i.evento_id = %s AND i.status = 'enviada'
            WHERE ac.id_associacao = %s
            ORDER BY ac.nome
        """, (evento_id, evento_id, id_assoc))
        academias = cur.fetchall()

        # Buscar todas as inscrições com dados do aluno e categoria
        cur.execute("""
            SELECT i.id, i.academia_id, i.aluno_id, i.dados_form, 
                   ac.nome as academia_nome, a.nome as aluno_nome
            FROM eventos_competicoes_inscricoes i
            INNER JOIN academias ac ON ac.id = i.academia_id
            LEFT JOIN alunos a ON a.id = i.aluno_id
            WHERE i.evento_id = %s AND i.status = 'enviada' AND ac.id_associacao = %s
            ORDER BY ac.nome, a.nome
        """, (evento_id, id_assoc))
        inscricoes = cur.fetchall()

        # Agrupar inscrições por academia e categoria
        academias_com_inscricoes = {}
        for ac in academias:
            academias_com_inscricoes[ac["academia_id"]] = {
                "academia_nome": ac["academia_nome"],
                "categorias": {}
            }

        for insc in inscricoes:
            try:
                academia_id = insc.get("academia_id")
                if not academia_id:
                    continue
                    
                dados_form_str = insc.get("dados_form")
                if dados_form_str:
                    try:
                        if isinstance(dados_form_str, str):
                            dados = json.loads(dados_form_str)
                        else:
                            dados = dados_form_str
                    except (json.JSONDecodeError, TypeError):
                        dados = {}
                else:
                    dados = {}
                    
                categoria = dados.get("categoria") or "Sem categoria"
                
                if academia_id not in academias_com_inscricoes:
                    continue
                
                if categoria not in academias_com_inscricoes[academia_id]["categorias"]:
                    academias_com_inscricoes[academia_id]["categorias"][categoria] = []
                
                academias_com_inscricoes[academia_id]["categorias"][categoria].append({
                    "id": insc.get("id"),
                    "aluno_id": insc.get("aluno_id"),
                    "aluno_nome": insc.get("aluno_nome") or "Avulso",
                    "dados_form": dados
                })
            except Exception as e:
                import logging
                logging.error(f"Erro ao processar inscrição {insc.get('id')}: {e}")
                continue

        # Calcular totais
        for ac_id, ac_data in academias_com_inscricoes.items():
            try:
                total_por_categoria = {cat: len(alunos) for cat, alunos in ac_data.get("categorias", {}).items()}
                ac_data["total_inscritos"] = sum(total_por_categoria.values())
                ac_data["total_por_categoria"] = total_por_categoria
            except Exception as e:
                import logging
                logging.error(f"Erro ao calcular totais para academia {ac_id}: {e}")
                ac_data["total_inscritos"] = 0
                ac_data["total_por_categoria"] = {}

        # Buscar campos do formulário
        campos_form = []
        if ev.get("id_formulario"):
            try:
                cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
                    (ev["id_formulario"],))
                campos_form = cur.fetchall()
            except Exception as e:
                import logging
                logging.error(f"Erro ao buscar campos_form: {e}")
                campos_form = []
        
        # Buscar configuração de exportação salva (se o campo existir)
        configuracao_exportacao = None
        try:
            cur.execute("SELECT configuracao_exportacao FROM eventos_competicoes WHERE id = %s", (evento_id,))
            ev_row = cur.fetchone()
            if ev_row and ev_row.get("configuracao_exportacao"):
                try:
                    config_str = ev_row["configuracao_exportacao"]
                    if isinstance(config_str, str) and config_str.strip():
                        configuracao_exportacao = json.loads(config_str)
                    elif isinstance(config_str, dict):
                        configuracao_exportacao = config_str
                except (json.JSONDecodeError, TypeError) as e:
                    # Log do erro mas continua sem configuração
                    import logging
                    logging.warning(f"Erro ao parsear configuracao_exportacao: {e}")
                    pass
        except Exception as e:
            # Campo não existe ainda ou erro na query, usar None
            import logging
            logging.warning(f"Erro ao buscar configuracao_exportacao: {e}")
            pass

        # Preparar agrupamento por categoria também
        categorias_com_inscricoes = {}
        for ac_id, ac_data in academias_com_inscricoes.items():
            for categoria, alunos in ac_data.get("categorias", {}).items():
                if categoria not in categorias_com_inscricoes:
                    categorias_com_inscricoes[categoria] = {
                        "academias": {},
                        "total_inscritos": 0
                    }
                # Agrupar alunos por academia dentro da categoria
                for aluno in alunos:
                    academia_nome = ac_data.get("academia_nome", "Academia Desconhecida")
                    if academia_nome not in categorias_com_inscricoes[categoria]["academias"]:
                        categorias_com_inscricoes[categoria]["academias"][academia_nome] = []
                    categorias_com_inscricoes[categoria]["academias"][academia_nome].append(aluno)
        
        # Calcular totais por categoria
        for categoria, cat_data in categorias_com_inscricoes.items():
            total = sum(len(alunos) for alunos in cat_data["academias"].values())
            cat_data["total_inscritos"] = total
        
        # Calcular total geral de inscritos
        total_geral_inscritos = len(inscricoes)
        
        # Buscar mapeamento de academias e graduações para exibir nomes ao invés de IDs
        academias_map = {}
        graduacoes_map = {}
        try:
            # Buscar todas as academias da associação
            cur.execute("SELECT id, nome FROM academias WHERE id_associacao = %s", (id_assoc,))
            for row in cur.fetchall():
                academias_map[str(row["id"])] = row["nome"]
            
            # Buscar todas as graduações
            cur.execute("SELECT id, faixa, graduacao, categoria FROM graduacao")
            for row in cur.fetchall():
                graduacao_nome = f"{row.get('faixa', '')} {row.get('graduacao', '')} {row.get('categoria', '')}".strip()
                graduacoes_map[str(row["id"])] = graduacao_nome if graduacao_nome else f"ID {row['id']}"
        except Exception as e:
            import logging
            logging.warning(f"Erro ao buscar mapeamentos de academias/graduações: {e}")
        
        # Obter tipo de agrupamento (padrão: academia)
        agrupamento = request.args.get("agrupamento", "academia")
        
        try:
            return render_template("eventos_competicoes/consolidar.html",
                evento=ev, academias_com_inscricoes=academias_com_inscricoes, 
                categorias_com_inscricoes=categorias_com_inscricoes,
                campos_form=campos_form, agrupamento=agrupamento,
                configuracao_exportacao=configuracao_exportacao or {},
                total_geral_inscritos=total_geral_inscritos,
                academias_map=academias_map,
                graduacoes_map=graduacoes_map,
                back_url=url_for("eventos_competicoes.lista"))
        except Exception as e:
            import logging
            logging.error(f"Erro ao renderizar template consolidar: {e}", exc_info=True)
            flash(f"Erro ao carregar página: {str(e)}", "danger")
            return redirect(url_for("eventos_competicoes.lista"))
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/anexo/<int:anexo_id>/download")
@login_required
def download_anexo(anexo_id):
    """Download de anexo de evento/competição."""
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    
    try:
        # Buscar anexo
        cur.execute("""
            SELECT eca.*, ec.id_associacao
            FROM eventos_competicoes_anexos eca
            INNER JOIN eventos_competicoes ec ON ec.id = eca.evento_id
            WHERE eca.id = %s
        """, (anexo_id,))
        anexo = cur.fetchone()
        
        if not anexo:
            flash("Anexo não encontrado.", "danger")
            return redirect(url_for("painel.home"))
        
        # Verificar permissão: associação (criador), academia (aderiu) ou aluno/responsável (aderiu)
        pode_baixar = False
        id_assoc = getattr(current_user, "id_associacao", None)
        
        if current_user.has_role("gestor_associacao") or current_user.has_role("admin"):
            # Associação pode baixar seus próprios eventos
            pode_baixar = (id_assoc == anexo["id_associacao"])
        elif current_user.has_role("gestor_academia") or current_user.has_role("professor"):
            # Academia pode baixar se aderiu ao evento
            academia_id = getattr(current_user, "id_academia", None)
            if academia_id:
                cur.execute("""
                    SELECT 1 FROM eventos_competicoes_adesao
                    WHERE evento_id = %s AND academia_id = %s AND aderiu = 1
                """, (anexo["evento_id"], academia_id))
                pode_baixar = cur.fetchone() is not None
        elif current_user.has_role("aluno") or current_user.has_role("responsavel"):
            # Aluno/responsável pode baixar se a academia aderiu
            if current_user.has_role("aluno"):
                cur.execute("SELECT id_academia FROM alunos WHERE usuario_id = %s LIMIT 1", (current_user.id,))
            else:
                aluno_id = request.args.get("aluno_id", type=int)
                if aluno_id:
                    cur.execute("SELECT id_academia FROM alunos WHERE id = %s", (aluno_id,))
                else:
                    cur.execute("""
                        SELECT a.id_academia FROM alunos a
                        INNER JOIN responsavel_alunos ra ON ra.aluno_id = a.id AND ra.usuario_id = %s
                        LIMIT 1
                    """, (current_user.id,))
            aluno_row = cur.fetchone()
            if aluno_row and aluno_row.get("id_academia"):
                academia_id = aluno_row["id_academia"]
                cur.execute("""
                    SELECT 1 FROM eventos_competicoes_adesao
                    WHERE evento_id = %s AND academia_id = %s AND aderiu = 1
                """, (anexo["evento_id"], academia_id))
                pode_baixar = cur.fetchone() is not None
        
        if not pode_baixar:
            flash("Você não tem permissão para baixar este anexo.", "danger")
            return redirect(url_for("painel.home"))
        
        # Caminho do arquivo
        filepath = os.path.join(current_app.root_path, "static", "uploads", UPLOAD_ANEXOS, anexo["caminho_arquivo"])
        
        if not os.path.exists(filepath):
            flash("Arquivo não encontrado no servidor.", "danger")
            return redirect(url_for("painel.home"))
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=anexo["nome_arquivo"],
            mimetype=anexo.get("tipo_mime") or "application/octet-stream"
        )
        
    except Exception as e:
        current_app.logger.error(f"Erro ao baixar anexo: {e}", exc_info=True)
        flash("Erro ao baixar anexo.", "danger")
        return redirect(url_for("painel.home"))
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/disponiveis")
@login_required
def disponiveis():
    """Aluno ou responsável: lista eventos (ativos e finalizados) com filtros."""
    if not (current_user.has_role("aluno") or current_user.has_role("responsavel")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))

    # Filtros
    filtro_status = request.args.get("status", "")  # ativo, finalizado, ou vazio (todos)
    filtro_mes = request.args.get("mes", "")  # 1-12 ou vazio (todos)
    filtro_ano = request.args.get("ano", "")  # ex: 2026 ou vazio

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True, buffered=True)
    try:
        academia_id = None
        if current_user.has_role("aluno"):
            cur.execute("SELECT id_academia, id FROM alunos WHERE usuario_id = %s LIMIT 1", (current_user.id,))
            r = cur.fetchone()
            academia_id = r["id_academia"] if r else None
        elif current_user.has_role("responsavel"):
            aluno_id = request.args.get("aluno_id", type=int)
            if aluno_id:
                cur.execute("SELECT 1 FROM responsavel_alunos WHERE usuario_id=%s AND aluno_id=%s", (current_user.id, aluno_id))
                resultado = cur.fetchone()
                # Garantir que o resultado foi totalmente consumido
                cur.fetchall()  # Consumir qualquer resultado restante
                if resultado:
                    cur.execute("SELECT id_academia FROM alunos WHERE id = %s", (aluno_id,))
                    r = cur.fetchone()
                    # Garantir que o resultado foi totalmente consumido
                    cur.fetchall()  # Consumir qualquer resultado restante
                    academia_id = r["id_academia"] if r else None
            if not academia_id:
                cur.execute("SELECT a.id_academia FROM alunos a INNER JOIN responsavel_alunos ra ON ra.aluno_id = a.id AND ra.usuario_id = %s LIMIT 1", (current_user.id,))
                r = cur.fetchone()
                # Garantir que o resultado foi totalmente consumido
                cur.fetchall()  # Consumir qualquer resultado restante
                academia_id = r["id_academia"] if r else None

        if not academia_id:
            flash("Nenhum aluno vinculado.", "warning")
            return redirect(url_for("painel.home"))

        # Verificar se colunas de taxa existem
        tem_coluna_taxa, tem_coluna_valor_taxa = _colunas_taxa_existem(cur)
        
        # Query base: todos os eventos que a academia aderiu
        if tem_coluna_taxa:
            query = """
                SELECT ec.id, ec.nome, ec.descricao, ec.tipo, ec.data_inicio, ec.data_fim, ec.status,
                       COALESCE(ec.tem_taxa, 0) as tem_taxa, ec.valor_taxa_sugerido,
                       ea.valor_taxa
                FROM eventos_competicoes ec
                INNER JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s AND ea.aderiu = 1
                WHERE 1=1
            """
        else:
            query = """
                SELECT ec.id, ec.nome, ec.descricao, ec.tipo, ec.data_inicio, ec.data_fim, ec.status,
                       0 as tem_taxa, NULL as valor_taxa_sugerido, NULL as valor_taxa
                FROM eventos_competicoes ec
                INNER JOIN eventos_competicoes_adesao ea ON ea.evento_id = ec.id AND ea.academia_id = %s AND ea.aderiu = 1
                WHERE 1=1
            """
        params = [academia_id]

        # Filtro de status
        if filtro_status == "ativo":
            query += " AND (ec.status = 'ativo' AND ec.data_fim > NOW())"
        elif filtro_status == "finalizado":
            query += " AND (ec.status = 'finalizado' OR ec.data_fim <= NOW())"
        # Se vazio, mostra todos

        # Filtro de mês
        if filtro_mes:
            try:
                mes_int = int(filtro_mes)
                if 1 <= mes_int <= 12:
                    query += " AND MONTH(ec.data_fim) = %s"
                    params.append(mes_int)
            except ValueError:
                pass

        # Filtro de ano
        if filtro_ano:
            try:
                ano_int = int(filtro_ano)
                if ano_int >= 2020:
                    query += " AND YEAR(ec.data_fim) = %s"
                    params.append(ano_int)
            except ValueError:
                pass

        query += " ORDER BY ec.data_fim DESC"
        cur.execute(query, tuple(params))
        eventos = cur.fetchall()
        
        # Buscar anexos para cada evento
        for ev in eventos:
            cur.execute("""
                SELECT id, nome_arquivo, tamanho_bytes, descricao
                FROM eventos_competicoes_anexos
                WHERE evento_id = %s
                ORDER BY created_at DESC
            """, (ev["id"],))
            ev["anexos"] = cur.fetchall()
            # Formatar tamanhos
            for anexo in ev["anexos"]:
                tamanho = anexo.get("tamanho_bytes") or 0
                if tamanho < 1024:
                    anexo["tamanho_formatado"] = f"{tamanho} B"
                elif tamanho < 1024 * 1024:
                    anexo["tamanho_formatado"] = f"{tamanho / 1024:.1f} KB"
                else:
                    anexo["tamanho_formatado"] = f"{tamanho / (1024 * 1024):.1f} MB"

        # Obter alunos vinculados
        if current_user.has_role("aluno"):
            cur.execute("SELECT id, nome FROM alunos WHERE usuario_id = %s", (current_user.id,))
            alunos_resp = cur.fetchall()
        else:
            cur.execute("SELECT a.id, a.nome FROM alunos a INNER JOIN responsavel_alunos ra ON ra.aluno_id = a.id AND ra.usuario_id = %s ORDER BY a.nome", (current_user.id,))
            alunos_resp = cur.fetchall()

        # Determinar aluno_id atual
        aluno_id_atual = request.args.get("aluno_id", type=int)
        if not aluno_id_atual and alunos_resp:
            aluno_id_atual = alunos_resp[0]["id"]

        # Verificar se tabela de pagamentos existe
        try:
            cur.execute("SHOW TABLES LIKE 'eventos_competicoes_inscricoes_pagamentos'")
            tabela_pagamentos_existe = cur.fetchone() is not None
        except Exception:
            tabela_pagamentos_existe = False
        
        # Marcar status visual e verificar inscrição
        for ev in eventos:
            ev["encerrado"] = ev.get("status") == "finalizado" or (ev.get("data_fim") and ev["data_fim"] < datetime.now())
            ev["ja_inscrito"] = False
            ev["status_pagamento"] = None  # None, 'pago', 'pendente', 'sem_registro'
            ev["total_pagamentos"] = 0
            ev["pagamentos_pendentes"] = 0
            ev["pagamentos_pagos"] = 0
            ev["valor_total_a_pagar"] = 0
            # Garantir que tem_taxa seja um inteiro
            if ev.get("tem_taxa"):
                ev["tem_taxa"] = int(ev["tem_taxa"])
            else:
                ev["tem_taxa"] = 0
            if aluno_id_atual:
                # Buscar todas as inscrições do aluno neste evento com status e pagamento
                if tabela_pagamentos_existe:
                    cur.execute("""
                        SELECT i.id, i.status, p.id as pagamento_id, 
                               COALESCE(p.pago_academia, 0) as pago_academia,
                               COALESCE(p.pago_associacao, 0) as pago_associacao,
                               p.valor
                        FROM eventos_competicoes_inscricoes i
                        LEFT JOIN eventos_competicoes_inscricoes_pagamentos p ON p.inscricao_id = i.id
                        WHERE i.evento_id = %s AND i.academia_id = %s AND i.aluno_id = %s
                        ORDER BY i.created_at DESC
                    """, (ev["id"], academia_id, aluno_id_atual))
                else:
                    cur.execute("""
                        SELECT i.id, i.status, NULL as pagamento_id, 
                               0 as pago_academia, 0 as pago_associacao, NULL as valor
                        FROM eventos_competicoes_inscricoes i
                        WHERE i.evento_id = %s AND i.academia_id = %s AND i.aluno_id = %s
                        ORDER BY i.created_at DESC
                    """, (ev["id"], academia_id, aluno_id_atual))
                inscricoes = cur.fetchall()
                
                # Verificar se há inscrições
                tem_inscricoes = len(inscricoes) > 0
                
                # Considerar "inscrito" se houver qualquer inscrição (independente do status)
                ev["ja_inscrito"] = tem_inscricoes
                
                # Status de confirmação separado (para mostrar se está confirmada pela associação)
                ev["confirmada_associacao"] = False
                if tem_inscricoes:
                    # Verificar se há pelo menos uma inscrição confirmada pela associação
                    for insc in inscricoes:
                        status_insc = insc.get("status", "")
                        pago_assoc = int(insc.get("pago_associacao", 0) or 0)
                        if status_insc == "confirmada" and pago_assoc == 1:
                            ev["confirmada_associacao"] = True
                            break
                
                # Se há inscrições e evento tem taxa, verificar status de pagamento
                if tem_inscricoes and ev["tem_taxa"] == 1 and tabela_pagamentos_existe:
                    ev["total_pagamentos"] = len([p for p in inscricoes if p.get("pagamento_id")])
                    
                    # Contar pagamentos por status
                    pagamentos_pendentes = 0  # pago_academia = 0
                    pagamentos_pagos = 0  # pago_academia = 1 mas não enviado ou não confirmado pela associação
                    pagamentos_enviados = 0  # status = 'enviada' e pago_associacao = 0
                    pagamentos_confirmados = 0  # status = 'confirmada' e pago_associacao = 1
                    
                    for insc in inscricoes:
                        if not insc.get("pagamento_id"):
                            continue
                        status_insc = insc.get("status", "")
                        pago_acad = int(insc.get("pago_academia", 0) or 0)
                        pago_assoc = int(insc.get("pago_associacao", 0) or 0)
                        
                        if status_insc == "confirmada" and pago_assoc == 1:
                            pagamentos_confirmados += 1
                        elif status_insc == "enviada" and pago_assoc == 0:
                            pagamentos_enviados += 1
                        elif pago_acad == 1:
                            pagamentos_pagos += 1
                        else:
                            pagamentos_pendentes += 1
                    
                    ev["pagamentos_pendentes"] = pagamentos_pendentes
                    ev["pagamentos_pagos"] = pagamentos_pagos
                    ev["pagamentos_enviados"] = pagamentos_enviados
                    ev["pagamentos_confirmados"] = pagamentos_confirmados
                    
                    # Calcular valor total a pagar (valor da taxa * quantidade de pagamentos pendentes)
                    valor_taxa = ev.get("valor_taxa") or ev.get("valor_taxa_sugerido") or 0
                    if valor_taxa and pagamentos_pendentes > 0:
                        ev["valor_total_a_pagar"] = float(valor_taxa) * pagamentos_pendentes
                    else:
                        ev["valor_total_a_pagar"] = 0
                    
                    # Determinar status principal de pagamento
                    if ev["total_pagamentos"] == 0:
                        ev["status_pagamento"] = "sem_registro"
                    elif pagamentos_confirmados > 0:
                        # Se há pelo menos uma confirmada, mostrar como confirmada
                        ev["status_pagamento"] = "confirmada"
                    elif pagamentos_enviados > 0:
                        # Se há pelo menos uma enviada mas não confirmada
                        ev["status_pagamento"] = "enviada"
                    elif pagamentos_pendentes > 0:
                        # Se há pendentes
                        ev["status_pagamento"] = "pendente"
                    elif pagamentos_pagos > 0:
                        # Se todas estão pagas mas não enviadas
                        ev["status_pagamento"] = "pago"
                    else:
                        ev["status_pagamento"] = "pendente"
                elif tem_inscricoes and ev["tem_taxa"] == 1:
                    # Evento tem taxa mas tabela de pagamentos não existe
                    ev["status_pagamento"] = "sem_registro"
                elif tem_inscricoes:
                    # Há inscrições mas evento não tem taxa - só considerar inscrito se status for 'confirmada'
                    ev["ja_inscrito"] = any(insc.get("status") == "confirmada" for insc in inscricoes)

        back_url = url_for("painel_responsavel.meu_perfil") if current_user.has_role("responsavel") else url_for("painel_aluno.painel")

        # Anos disponíveis para filtro
        cur.execute("SELECT DISTINCT YEAR(data_fim) as ano FROM eventos_competicoes ORDER BY ano DESC")
        anos_disponiveis = [r["ano"] for r in cur.fetchall() if r["ano"]]

        return render_template("eventos_competicoes/disponiveis.html",
            eventos=eventos, academia_id=academia_id, alunos=alunos_resp,
            filtro_status=filtro_status, filtro_mes=filtro_mes, filtro_ano=filtro_ano,
            anos_disponiveis=anos_disponiveis,
            back_url=back_url)
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/<int:evento_id>/configuracao-exportacao", methods=["GET"])
@login_required
def get_configuracao_exportacao(evento_id):
    """Retorna configuração de exportação salva."""
    if session.get("modo_painel") != "associacao" or not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        return jsonify({"error": "Acesso negado"}), 403
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        return jsonify({"error": "Associação não encontrada"}), 404
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT configuracao_exportacao FROM eventos_competicoes WHERE id = %s AND id_associacao = %s",
            (evento_id, id_assoc))
        ev = cur.fetchone()
        if not ev:
            return jsonify({"error": "Evento não encontrado"}), 404
        
        config = None
        if ev.get("configuracao_exportacao"):
            try:
                config = json.loads(ev["configuracao_exportacao"]) if isinstance(ev["configuracao_exportacao"], str) else ev["configuracao_exportacao"]
            except Exception:
                pass
        
        return jsonify({"campos": config.get("campos", []) if config else []})
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/<int:evento_id>/configuracao-exportacao", methods=["POST"])
@login_required
def salvar_configuracao_exportacao(evento_id):
    """Salva configuração de exportação."""
    if session.get("modo_painel") != "associacao" or not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        return jsonify({"success": False, "error": "Acesso negado"}), 403
    
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        return jsonify({"success": False, "error": "Associação não encontrada"}), 404
    
    data = request.get_json()
    campos = data.get("campos", [])
    
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar se evento existe e pertence à associação
        cur.execute("SELECT id FROM eventos_competicoes WHERE id = %s AND id_associacao = %s",
            (evento_id, id_assoc))
        if not cur.fetchone():
            return jsonify({"success": False, "error": "Evento não encontrado"}), 404
        
        config = json.dumps({"campos": campos}, ensure_ascii=False)
        cur.execute("UPDATE eventos_competicoes SET configuracao_exportacao = %s WHERE id = %s",
            (config, evento_id))
        conn.commit()
        
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/<int:evento_id>/configurar-impressao")
@login_required
def configurar_impressao(evento_id):
    """Página de configuração de impressão/exportação com opções de escala e layout."""
    try:
        if session.get("modo_painel") != "associacao" or not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
            flash("Acesso negado.", "danger")
            return redirect(url_for("painel.home"))
        id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
        if not id_assoc:
            return redirect(url_for("associacao.gerenciamento_associacao"))

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        try:
            cur.execute("""
                SELECT ec.id, ec.nome, ec.data_fim, ec.id_formulario, ec.configuracao_exportacao
                FROM eventos_competicoes ec WHERE ec.id = %s AND ec.id_associacao = %s
            """, (evento_id, id_assoc))
            ev = cur.fetchone()
            if not ev:
                flash("Evento não encontrado.", "danger")
                return redirect(url_for("eventos_competicoes.lista"))

            # Buscar campos do formulário
            campos_form = []
            if ev.get("id_formulario"):
                try:
                    cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
                        (ev["id_formulario"],))
                    campos_form = cur.fetchall()
                except Exception as e:
                    import logging
                    logging.error(f"Erro ao buscar campos_form: {e}")
                    campos_form = []
            
            # Buscar configuração de exportação salva
            configuracao_exportacao = None
            try:
                if ev.get("configuracao_exportacao"):
                    try:
                        config_str = ev["configuracao_exportacao"]
                        if isinstance(config_str, str) and config_str.strip():
                            configuracao_exportacao = json.loads(config_str)
                        elif isinstance(config_str, dict):
                            configuracao_exportacao = config_str
                    except (json.JSONDecodeError, TypeError) as e:
                        import logging
                        logging.warning(f"Erro ao parsear configuracao_exportacao: {e}")
                        pass
            except Exception as e:
                import logging
                logging.warning(f"Erro ao buscar configuracao_exportacao: {e}")
                pass

            # Obter tipo de agrupamento (padrão: academia)
            agrupamento = request.args.get("agrupamento", "academia")
            
            # Buscar algumas inscrições para preview (limitar a 10 para performance)
            cur.execute("""
                SELECT i.id, i.dados_form, ac.nome as academia_nome, a.nome as aluno_nome
                FROM eventos_competicoes_inscricoes i
                INNER JOIN academias ac ON ac.id = i.academia_id
                LEFT JOIN alunos a ON a.id = i.aluno_id
                WHERE i.evento_id = %s AND i.status = 'enviada' AND ac.id_associacao = %s
                ORDER BY ac.nome, a.nome
                LIMIT 10
            """, (evento_id, id_assoc))
            preview_inscricoes_raw = cur.fetchall()
            
            # Processar dados_form de JSON string para dict
            preview_inscricoes = []
            for insc in preview_inscricoes_raw:
                dados_form_parsed = {}
                if insc.get("dados_form"):
                    try:
                        if isinstance(insc["dados_form"], str):
                            dados_form_parsed = json.loads(insc["dados_form"])
                        elif isinstance(insc["dados_form"], dict):
                            dados_form_parsed = insc["dados_form"]
                    except (json.JSONDecodeError, TypeError):
                        dados_form_parsed = {}
                preview_inscricoes.append({
                    "id": insc.get("id"),
                    "dados_form": dados_form_parsed,
                    "academia_nome": insc.get("academia_nome"),
                    "aluno_nome": insc.get("aluno_nome")
                })
            
            return render_template("eventos_competicoes/configurar_impressao.html",
                evento=ev, campos_form=campos_form, agrupamento=agrupamento,
                configuracao_exportacao=configuracao_exportacao or {},
                preview_inscricoes=preview_inscricoes,
                back_url=url_for("eventos_competicoes.consolidar", evento_id=evento_id))
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        import logging
        logging.error(f"Erro em configurar_impressao: {e}", exc_info=True)
        flash(f"Erro ao carregar página de configuração: {str(e)}", "danger")
        return redirect(url_for("eventos_competicoes.consolidar", evento_id=evento_id))


@bp_eventos_competicoes.route("/<int:evento_id>/imprimir")
@login_required
def imprimir(evento_id):
    """Página de impressão otimizada com configurações de escala."""
    if session.get("modo_painel") != "associacao" or not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        return redirect(url_for("associacao.gerenciamento_associacao"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT ec.id, ec.nome, ec.data_fim, ec.id_formulario
            FROM eventos_competicoes ec WHERE ec.id = %s AND ec.id_associacao = %s
        """, (evento_id, id_assoc))
        ev = cur.fetchone()
        if not ev:
            flash("Evento não encontrado.", "danger")
            return redirect(url_for("eventos_competicoes.lista"))

        # Obter campos selecionados e configurações
        campos_selecionados = request.args.getlist("campos")
        agrupamento = request.args.get("agrupamento", "academia")
        escala = request.args.get("escala", "100", type=int)
        tamanho_fonte = request.args.get("tamanho_fonte", "10", type=int)
        orientacao = request.args.get("orientacao", "paisagem")
        margem_vertical = request.args.get("margem_vertical", "20", type=int)
        margem_horizontal = request.args.get("margem_horizontal", "20", type=int)
        
        # Obter parâmetros de ordenação
        ordenar_por = request.args.get("ordenar_por", "")
        ordenar_direcao = request.args.get("ordenar_direcao", "asc").upper()
        
        # Buscar campos do formulário para validação de ordenação
        campos_form = []
        if ev.get("id_formulario"):
            try:
                cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
                    (ev["id_formulario"],))
                campos_form = cur.fetchall()
            except Exception:
                campos_form = []

        # Buscar academias e inscrições (mesma lógica do consolidar)
        cur.execute("""
            SELECT DISTINCT ac.id as academia_id, ac.nome as academia_nome
            FROM academias ac
            INNER JOIN eventos_competicoes_adesao ea ON ea.academia_id = ac.id AND ea.evento_id = %s AND ea.aderiu = 1
            INNER JOIN eventos_competicoes_inscricoes i ON i.academia_id = ac.id AND i.evento_id = %s AND i.status = 'enviada'
            WHERE ac.id_associacao = %s
            ORDER BY ac.nome
        """, (evento_id, evento_id, id_assoc))
        academias = cur.fetchall()
        
        cur.execute("""
            SELECT i.id, i.academia_id, i.aluno_id, i.dados_form, 
                   ac.nome as academia_nome, a.nome as aluno_nome
            FROM eventos_competicoes_inscricoes i
            INNER JOIN academias ac ON ac.id = i.academia_id
            LEFT JOIN alunos a ON a.id = i.aluno_id
            WHERE i.evento_id = %s AND i.status = 'enviada' AND ac.id_associacao = %s
            ORDER BY ac.nome, a.nome
        """, (evento_id, id_assoc))
        inscricoes = cur.fetchall()
        
        # Aplicar ordenação por campo do formulário se especificado
        if ordenar_por and ordenar_por in {c["campo_chave"] for c in campos_form}:
            def get_sort_value(row):
                dados_form_str = row.get("dados_form")
                if dados_form_str:
                    try:
                        if isinstance(dados_form_str, str):
                            dados = json.loads(dados_form_str)
                        else:
                            dados = dados_form_str
                    except (json.JSONDecodeError, TypeError):
                        dados = {}
                else:
                    dados = {}
                valor = dados.get(ordenar_por, "")
                # Tentar converter para número se possível
                try:
                    if isinstance(valor, (int, float)):
                        return (0, valor)
                    valor_str = str(valor).strip()
                    if valor_str.replace('.', '').replace('-', '').isdigit():
                        return (0, float(valor_str))
                except (ValueError, TypeError):
                    pass
                return (1, str(valor).lower())
            
            inscricoes.sort(key=get_sort_value, reverse=(ordenar_direcao == "DESC"))

        # Agrupar inscrições (mesma lógica do consolidar)
        academias_com_inscricoes = {}
        for ac in academias:
            academias_com_inscricoes[ac["academia_id"]] = {
                "academia_nome": ac["academia_nome"],
                "categorias": {}
            }

        for insc in inscricoes:
            try:
                academia_id = insc.get("academia_id")
                if not academia_id:
                    continue
                    
                dados_form_str = insc.get("dados_form")
                if dados_form_str:
                    try:
                        if isinstance(dados_form_str, str):
                            dados = json.loads(dados_form_str)
                        else:
                            dados = dados_form_str
                    except (json.JSONDecodeError, TypeError):
                        dados = {}
                else:
                    dados = {}
                    
                categoria = dados.get("categoria") or "Sem categoria"
                
                if academia_id not in academias_com_inscricoes:
                    continue
                
                if categoria not in academias_com_inscricoes[academia_id]["categorias"]:
                    academias_com_inscricoes[academia_id]["categorias"][categoria] = []
                
                academias_com_inscricoes[academia_id]["categorias"][categoria].append({
                    "id": insc.get("id"),
                    "aluno_id": insc.get("aluno_id"),
                    "aluno_nome": insc.get("aluno_nome") or "Avulso",
                    "dados_form": dados
                })
            except Exception as e:
                import logging
                logging.error(f"Erro ao processar inscrição {insc.get('id')}: {e}")
                continue

        # Calcular totais
        for ac_id, ac_data in academias_com_inscricoes.items():
            try:
                total_por_categoria = {cat: len(alunos) for cat, alunos in ac_data.get("categorias", {}).items()}
                ac_data["total_inscritos"] = sum(total_por_categoria.values())
                ac_data["total_por_categoria"] = total_por_categoria
            except Exception:
                ac_data["total_inscritos"] = 0
                ac_data["total_por_categoria"] = {}

        # Preparar agrupamento por categoria também
        categorias_com_inscricoes = {}
        for ac_id, ac_data in academias_com_inscricoes.items():
            for categoria, alunos in ac_data.get("categorias", {}).items():
                if categoria not in categorias_com_inscricoes:
                    categorias_com_inscricoes[categoria] = {
                        "academias": {},
                        "total_inscritos": 0
                    }
                for aluno in alunos:
                    academia_nome = ac_data.get("academia_nome", "Academia Desconhecida")
                    if academia_nome not in categorias_com_inscricoes[categoria]["academias"]:
                        categorias_com_inscricoes[categoria]["academias"][academia_nome] = []
                    categorias_com_inscricoes[categoria]["academias"][academia_nome].append(aluno)
        
        for categoria, cat_data in categorias_com_inscricoes.items():
            total = sum(len(alunos) for alunos in cat_data["academias"].values())
            cat_data["total_inscritos"] = total

        # Buscar campos do formulário
        campos_form = []
        if ev.get("id_formulario"):
            try:
                cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
                    (ev["id_formulario"],))
                campos_form = cur.fetchall()
            except Exception:
                campos_form = []
        
        # Filtrar campos se selecionados e manter ordem da URL
        campos_para_imprimir = []
        if campos_selecionados:
            # Criar um mapa para acesso rápido
            campos_map = {c["campo_chave"]: c for c in campos_form}
            # Manter a ordem dos campos_selecionados (que vem da URL na ordem correta)
            for campo_chave in campos_selecionados:
                if campo_chave in campos_map:
                    campos_para_imprimir.append(campos_map[campo_chave])
        else:
            campos_para_imprimir = campos_form
        
        # Preparar inscrições simples para impressão (sem agrupamento complexo)
        inscricoes_para_imprimir = []
        for insc in inscricoes:
            try:
                dados_form_str = insc.get("dados_form")
                if dados_form_str:
                    try:
                        if isinstance(dados_form_str, str):
                            dados = json.loads(dados_form_str)
                        else:
                            dados = dados_form_str
                    except (json.JSONDecodeError, TypeError):
                        dados = {}
                else:
                    dados = {}
                
                inscricoes_para_imprimir.append({
                    "id": insc.get("id"),
                    "aluno_nome": insc.get("aluno_nome") or "Avulso",
                    "academia_nome": insc.get("academia_nome") or "Academia Desconhecida",
                    "dados_form": dados
                })
            except Exception as e:
                import logging
                logging.error(f"Erro ao processar inscrição para impressão {insc.get('id')}: {e}")
                continue

        # Função helper para formatação de valores (mesma lógica do consolidar)
        def _formatar_valor(campo_chave, valor):
            if campo_chave == 'sexo':
                if valor == 'M':
                    return 'Masculino'
                elif valor == 'F':
                    return 'Feminino'
                return valor or '-'
            elif campo_chave in ['data_nascimento', 'ultimo_exame_faixa', 'rg_data_emissao', 'data_cadastro_zempo']:
                if valor:
                    try:
                        if isinstance(valor, str) and len(valor) >= 10:
                            if '-' in valor[:10]:
                                # Formato ISO: YYYY-MM-DD
                                return f"{valor[8:10]}/{valor[5:7]}/{valor[0:4]}"
                    except Exception:
                        pass
                return valor or '-'
            return valor or '-'

        return render_template("eventos_competicoes/imprimir.html",
            evento=ev, academias_com_inscricoes=academias_com_inscricoes,
            categorias_com_inscricoes=categorias_com_inscricoes,
            campos_form=campos_para_imprimir, agrupamento=agrupamento,
            escala=escala, tamanho_fonte=tamanho_fonte,
            orientacao=orientacao, margem_vertical=margem_vertical,
            margem_horizontal=margem_horizontal,
            inscricoes_para_imprimir=inscricoes_para_imprimir)
    finally:
        cur.close()
        conn.close()


@bp_eventos_competicoes.route("/<int:evento_id>/exportar")
@login_required
def exportar(evento_id):
    """Exporta inscrições em PDF ou Excel conforme formato."""
    fmt = request.args.get("formato", "excel").lower()
    campos_selecionados = request.args.getlist("campos")  # Lista de campos selecionados
    
    # Filtrar campos vazios ou inválidos
    campos_selecionados = [c for c in campos_selecionados if c and c.strip() and c.strip().lower() not in ("aluno", "academia")]
    
    if session.get("modo_painel") != "associacao" or not (current_user.has_role("gestor_associacao") or current_user.has_role("admin")):
        flash("Acesso negado.", "danger")
        return redirect(url_for("painel.home"))
    id_assoc = getattr(current_user, "id_associacao", None) or session.get("associacao_gerenciamento_id")
    if not id_assoc:
        return redirect(url_for("associacao.gerenciamento_associacao"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, nome, id_formulario, configuracao_exportacao FROM eventos_competicoes WHERE id = %s AND id_associacao = %s",
            (evento_id, id_assoc))
        ev = cur.fetchone()
        if not ev:
            return redirect(url_for("eventos_competicoes.lista"))

        cur.execute("SELECT campo_chave, label, ordem FROM formularios_campos WHERE formulario_id = %s ORDER BY ordem",
            (ev["id_formulario"] or 0,))
        campos_form = cur.fetchall()
        
        # Se campos foram selecionados via URL, usar eles; senão, usar configuração salva ou todos
        if campos_selecionados:
            # Filtrar campos_form pelos selecionados e manter ordem
            campos_para_exportar = []
            for chave in campos_selecionados:
                # Ignorar campos vazios ou inválidos
                if not chave or chave.strip() == "":
                    continue
                campo = next((c for c in campos_form if c["campo_chave"] == chave), None)
                if campo:
                    campos_para_exportar.append(campo)
                # Se campo não encontrado, não adicionar (evitar campos inválidos)
        else:
            # Tentar usar configuração salva
            config = None
            if ev.get("configuracao_exportacao"):
                try:
                    config = json.loads(ev["configuracao_exportacao"]) if isinstance(ev["configuracao_exportacao"], str) else ev["configuracao_exportacao"]
                except Exception:
                    pass
            
            if config and config.get("campos"):
                # Ordenar campos_form conforme configuração
                ordem_map = {c["chave"]: idx for idx, c in enumerate(config["campos"])}
                campos_para_exportar = sorted(
                    [c for c in campos_form if c["campo_chave"] in ordem_map],
                    key=lambda x: ordem_map.get(x["campo_chave"], 999)
                )
            else:
                # Usar todos os campos na ordem padrão
                campos_para_exportar = campos_form
        
        # Obter orientação do PDF (da configuração salva ou da URL, padrão: paisagem)
        orientacao_pdf = "paisagem"
        if campos_selecionados:
            # Se há campos na URL, tentar pegar orientação da URL também
            orientacao_url = request.args.get("orientacao", "").lower()
            if orientacao_url in ("paisagem", "retrato"):
                orientacao_pdf = orientacao_url
        elif config and config.get("orientacao_pdf"):
            orientacao_pdf = config.get("orientacao_pdf", "paisagem")
        
        # Garantir que labels e chaves correspondem APENAS aos campos selecionados
        # Remover qualquer campo que não esteja na lista de campos do formulário
        campos_validos_chaves = {c["campo_chave"] for c in campos_form}
        campos_para_exportar = [c for c in campos_para_exportar if c["campo_chave"] in campos_validos_chaves]
        
        labels = [c["label"] for c in campos_para_exportar]
        chaves = [c["campo_chave"] for c in campos_para_exportar]

        cur.execute("SELECT nome FROM associacoes WHERE id = %s", (id_assoc,))
        assoc_row = cur.fetchone()
        assoc_nome = assoc_row["nome"] if assoc_row else ""

        # Obter parâmetros de ordenação
        ordenar_por = request.args.get("ordenar_por", "")
        ordenar_direcao = request.args.get("ordenar_direcao", "asc").upper()
        
        # Construir ORDER BY dinâmico
        order_by = "ac.nome, a.nome"  # Padrão
        if ordenar_por:
            # Validar campo de ordenação (precisa estar nos campos do formulário)
            campos_validos = {c["campo_chave"] for c in campos_form}
            if ordenar_por in campos_validos:
                # Ordenação será feita após parsear JSON, então manteremos ordem padrão aqui
                # e ordenaremos depois no Python
                order_by = "ac.nome, a.nome"
            else:
                order_by = "ac.nome, a.nome"
        
        cur.execute("""
            SELECT i.id, i.aluno_id, i.dados_form, i.inclusao_avulsa, ac.nome as academia_nome, a.nome as aluno_nome
            FROM eventos_competicoes_inscricoes i
            INNER JOIN academias ac ON ac.id = i.academia_id
            LEFT JOIN alunos a ON a.id = i.aluno_id
            WHERE i.evento_id = %s AND i.status = 'enviada' AND ac.id_associacao = %s
            ORDER BY ac.nome, a.nome
        """, (evento_id, id_assoc))
        rows = cur.fetchall()
        
        # Aplicar ordenação por campo do formulário se especificado
        if ordenar_por and ordenar_por in {c["campo_chave"] for c in campos_form}:
            def get_sort_value(row):
                dados_form_str = row.get("dados_form")
                if dados_form_str:
                    try:
                        if isinstance(dados_form_str, str):
                            dados = json.loads(dados_form_str)
                        else:
                            dados = dados_form_str
                    except (json.JSONDecodeError, TypeError):
                        dados = {}
                else:
                    dados = {}
                valor = dados.get(ordenar_por, "")
                # Tentar converter para número se possível
                try:
                    if isinstance(valor, (int, float)):
                        return (0, valor)
                    valor_str = str(valor).strip()
                    if valor_str.replace('.', '').replace('-', '').isdigit():
                        return (0, float(valor_str))
                except (ValueError, TypeError):
                    pass
                return (1, str(valor).lower())
            
            rows.sort(key=get_sort_value, reverse=(ordenar_direcao == "DESC"))
    finally:
        cur.close()
        conn.close()

    def _formatar_valor(chave, val):
        """Formata valor: sexo M/F -> Masculino/Feminino; datas -> dd/mm/yyyy."""
        if val is None or val == "":
            return ""
        s = str(val).strip()
        if chave == "sexo":
            if s.upper() == "M":
                return "Masculino"
            if s.upper() == "F":
                return "Feminino"
            return s
        if chave in ("data_nascimento", "ultimo_exame_faixa", "rg_data_emissao", "data_cadastro_zempo"):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    d = datetime.strptime(s[:10], fmt)
                    return d.strftime("%d/%m/%Y")
                except (ValueError, TypeError):
                    continue
        return s

    if fmt == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            
            # Criar workbook e worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Inscrições"
            
            # Estilizar cabeçalho
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Montar header APENAS com os campos selecionados (incluindo coluna # como na prévia)
            # Adicionar coluna "#" no início
            cell_numero = ws.cell(row=1, column=1, value="#")
            cell_numero.fill = header_fill
            cell_numero.font = header_font
            cell_numero.alignment = header_alignment
            
            # Adicionar labels dos campos
            for col_idx, label in enumerate(labels, start=2):
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Ajustar largura das colunas automaticamente baseado no conteúdo
            max_widths = {}
            # Coluna "#" - largura fixa pequena
            col_letter_numero = ws.cell(row=1, column=1).column_letter
            max_widths[col_letter_numero] = 8  # Largura pequena para coluna "#"
            ws.column_dimensions[col_letter_numero].width = max_widths[col_letter_numero]
            
            # Ajustar largura das outras colunas
            for col_idx, campo in enumerate(campos_para_exportar, start=2):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                # Começar com largura do cabeçalho
                max_width = len(str(labels[col_idx - 2])) + 2
                # Verificar conteúdo das células para encontrar maior valor
                for r in rows:
                    dados = json.loads(r["dados_form"]) if r.get("dados_form") else {}
                    k = campo["campo_chave"]
                    valor = dados.get(k, "")
                    if valor == "" and k == "id_academia":
                        valor = r.get("academia_nome", "")
                    elif valor == "" and k == "nome":
                        valor = r.get("aluno_nome", "")
                    valor_formatado = _formatar_valor(k, valor)
                    max_width = max(max_width, len(str(valor_formatado)) + 2)
                # Limitar entre mínimo e máximo
                max_widths[col_letter] = min(max(max_width, 10), 50)
                ws.column_dimensions[col_letter].width = max_widths[col_letter]
            
            # Adicionar dados nas linhas
            for row_idx, r in enumerate(rows, start=2):
                dados = json.loads(r["dados_form"]) if r.get("dados_form") else {}
                
                # Adicionar número da linha na primeira coluna
                cell_numero = ws.cell(row=row_idx, column=1, value=str(row_idx - 1))
                cell_numero.alignment = Alignment(vertical="top", horizontal="center")
                
                # Adicionar APENAS os campos selecionados
                for col_idx, campo in enumerate(campos_para_exportar, start=2):
                    k = campo["campo_chave"]
                    # Se o campo não estiver em dados_form, tentar buscar de outras fontes apenas se necessário
                    valor = dados.get(k, "")
                    if valor == "" and k == "id_academia":
                        # Se for id_academia e não tiver valor, buscar nome da academia do join
                        valor = r.get("academia_nome", "")
                    elif valor == "" and k == "nome":
                        # Se for nome e não tiver valor, buscar nome do aluno do join
                        valor = r.get("aluno_nome", "")
                    
                    # Formatar valor
                    valor_formatado = _formatar_valor(k, valor)
                    
                    # Adicionar célula
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor_formatado)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Congelar primeira linha (cabeçalho)
            ws.freeze_panes = "A2"
            
            # Salvar em BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            resp = Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp.headers["Content-Disposition"] = f'attachment; filename="inscricoes_{ev["nome"][:30]}.xlsx"'
            return resp
        except ImportError:
            flash("Biblioteca openpyxl não instalada. Execute: pip install openpyxl", "warning")
            return redirect(url_for("eventos_competicoes.consolidar", evento_id=evento_id))
        except Exception as e:
            flash(f"Erro ao exportar: {e}", "danger")
            return redirect(url_for("eventos_competicoes.consolidar", evento_id=evento_id))

    # PDF: tabela única, orientação configurável (paisagem/retrato), evento no título, associação abaixo
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from io import BytesIO
        buffer = BytesIO()
        
        # Obter configurações de impressão da URL
        escala_pdf = request.args.get("escala", "100", type=int)
        tamanho_fonte_pdf = request.args.get("tamanho_fonte", "10", type=int)
        margem_vertical = request.args.get("margem_vertical", "20", type=int)
        margem_horizontal = request.args.get("margem_horizontal", "20", type=int)
        
        # Usar orientação configurada (paisagem ou retrato)
        if orientacao_pdf == "retrato":
            page_size = A4
        else:
            page_size = landscape(A4)
        
        # Aplicar margens configuradas (converter mm para pontos: 1mm = 2.83465 pontos)
        left_margin = max(10, margem_horizontal * 2.83465)
        right_margin = max(10, margem_horizontal * 2.83465)
        top_margin = max(10, margem_vertical * 2.83465)
        bottom_margin = max(10, margem_vertical * 2.83465)
        
        doc = SimpleDocTemplate(buffer, pagesize=page_size, 
                                leftMargin=left_margin, rightMargin=right_margin, 
                                topMargin=top_margin, bottomMargin=bottom_margin)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph(f"<b>{ev['nome']}</b>", styles["Title"]))
        elements.append(Paragraph(assoc_nome, styles["Heading3"]))
        elements.append(Spacer(1, 12))
        
        # Montar header APENAS com os campos selecionados (incluindo coluna # como na prévia)
        tbl_header = ["#"] + labels  # Adicionar coluna "#" no início como na prévia
        tbl_rows = [tbl_header]
        for idx, r in enumerate(rows, start=1):
            dados = json.loads(r["dados_form"]) if r.get("dados_form") else {}
            linha = [str(idx)]  # Adicionar número da linha no início
            
            # Adicionar APENAS os campos selecionados na ordem correta
            for campo in campos_para_exportar:
                k = campo["campo_chave"]
                # Se o campo não estiver em dados_form, tentar buscar de outras fontes apenas se necessário
                valor = dados.get(k, "")
                if valor == "" and k == "id_academia":
                    # Se for id_academia e não tiver valor, buscar nome da academia do join
                    valor = r.get("academia_nome", "")
                elif valor == "" and k == "nome":
                    # Se for nome e não tiver valor, buscar nome do aluno do join
                    valor = r.get("aluno_nome", "Avulso")
                linha.append(_formatar_valor(k, valor))
            tbl_rows.append(linha)
        if len(tbl_rows) > 1:
            n_cols = len(tbl_header)
            # Calcular largura disponível considerando margens
            available_width = page_size[0] - left_margin - right_margin
            
            # Aplicar escala às larguras das colunas
            escala_factor = escala_pdf / 100.0
            # A coluna "#" deve ser menor (30 pontos), as outras dividem o espaço restante
            col_width_numero = 30 * escala_factor  # Coluna "#"
            largura_restante = (available_width - col_width_numero) * escala_factor
            base_col_width = largura_restante / (n_cols - 1) if n_cols > 1 else largura_restante
            col_widths = [max(20, col_width_numero)] + [max(30, base_col_width)] * (n_cols - 1)
            
            t = Table(tbl_rows, repeatRows=1, colWidths=col_widths)
            t.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), tamanho_fonte_pdf),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e0e0")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph("Nenhuma inscrição enviada.", styles["Normal"]))
        doc.build(elements)
        buffer.seek(0)
        resp = Response(buffer.getvalue(), mimetype="application/pdf")
        resp.headers["Content-Disposition"] = f'attachment; filename="inscricoes_{ev["nome"][:30]}.pdf"'
        return resp
    except ImportError:
        flash("Biblioteca reportlab não instalada. Use exportar em Excel.", "warning")
        return redirect(url_for("eventos_competicoes.consolidar", evento_id=evento_id))
    except Exception as e:
        flash(f"Erro ao gerar PDF: {e}", "danger")
        return redirect(url_for("eventos_competicoes.consolidar", evento_id=evento_id))


@bp_eventos_competicoes.route("/buscar-categorias", methods=["POST"])
@login_required
def buscar_categorias():
    """Rota AJAX para buscar categorias disponíveis baseado em peso, idade e gênero."""
    try:
        peso = request.json.get("peso")
        data_nascimento = request.json.get("data_nascimento")
        genero = request.json.get("genero")
        
        if not peso or not data_nascimento or not genero:
            return Response(json.dumps({"categorias": []}), mimetype="application/json")
        
        # Calcular idade em ano civil
        try:
            from datetime import datetime as dt
            nasc = dt.strptime(data_nascimento[:10], "%Y-%m-%d").date()
            hoje = date.today()
            idade_ano_civil = hoje.year - nasc.year
        except Exception:
            return Response(json.dumps({"categorias": []}), mimetype="application/json")
        
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        try:
            genero_upper = genero.upper()
            if genero_upper not in ("M", "F"):
                return Response(json.dumps({"categorias": []}), mimetype="application/json")
            
            peso_float = float(peso)
            
            # Mapear M/F para MASCULINO/FEMININO
            genero_db = "MASCULINO" if genero_upper == "M" else "FEMININO" if genero_upper == "F" else genero_upper
            
            # Verificar se coluna id_classe existe
            try:
                cur.execute("SHOW COLUMNS FROM categorias LIKE 'id_classe'")
                tem_id_classe = cur.fetchone() is not None
            except Exception:
                tem_id_classe = False
            
            if tem_id_classe:
                cur.execute("""
                    SELECT id, categoria, nome_categoria, id_classe, peso_min, peso_max, idade_min, idade_max
                    FROM categorias
                    WHERE UPPER(genero) = UPPER(%s)
                    AND (
                        (idade_min IS NULL OR %s >= idade_min)
                        AND (idade_max IS NULL OR %s <= idade_max)
                    )
                    AND (
                        (peso_min IS NULL OR %s >= peso_min)
                        AND (peso_max IS NULL OR %s <= peso_max)
                    )
                    ORDER BY nome_categoria
                """, (genero_db, idade_ano_civil, idade_ano_civil, peso_float, peso_float))
            else:
                cur.execute("""
                    SELECT id, categoria, nome_categoria, NULL as id_classe, peso_min, peso_max, idade_min, idade_max
                    FROM categorias
                    WHERE UPPER(genero) = UPPER(%s)
                    AND (
                        (idade_min IS NULL OR %s >= idade_min)
                        AND (idade_max IS NULL OR %s <= idade_max)
                    )
                    AND (
                        (peso_min IS NULL OR %s >= peso_min)
                        AND (peso_max IS NULL OR %s <= peso_max)
                    )
                    ORDER BY nome_categoria
                """, (genero_db, idade_ano_civil, idade_ano_civil, peso_float, peso_float))
            
            categorias = cur.fetchall()
            return Response(json.dumps({"categorias": categorias}), mimetype="application/json")
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        return Response(json.dumps({"categorias": [], "erro": str(e)}), mimetype="application/json")
